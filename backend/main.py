# LLM chat/résumés : Gemini | PDF Vision : Claude Haiku (extract_worker.py)
import asyncio
import base64
import hashlib
import io
import json
import math
import os
import re
import subprocess
import sys
import threading
import time
import traceback
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import asynccontextmanager
from datetime import datetime, timezone, timedelta
from typing import Optional

import numpy as np
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from sentence_transformers import SentenceTransformer, CrossEncoder
from supabase import create_client as _make_sb_client, Client
import google.generativeai as genai
from google.generativeai.types import HarmCategory, HarmBlockThreshold
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError as _GHttpError
from googleapiclient.http import MediaIoBaseDownload

# Disable HuggingFace tokenizer's internal Rust thread pool — it conflicts with
# Python's ThreadPoolExecutor and corrupts glibc heap (SIGABRT / free() crash).
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")

# ── Model loading ─────────────────────────────────────────────────────────────
model: Optional[SentenceTransformer] = None
reranker: Optional[CrossEncoder] = None

# Single-worker executors: model.encode() / reranker.predict() must never run
# in two threads at once — concurrent inference corrupts the glibc heap (SIGABRT).
_EMBED_EXECUTOR = ThreadPoolExecutor(max_workers=1)
_RERANK_EXECUTOR = ThreadPoolExecutor(max_workers=1)

# In-memory sync state — lets the frontend poll progress after an SSE drop.
_sync_state: dict = {}  # key: f"{client_id}|{folder_id}"

# ── Langfuse — observabilité LLM (API v4), fail-open ──────────────────────
# Lit LANGFUSE_PUBLIC_KEY, LANGFUSE_SECRET_KEY, LANGFUSE_HOST depuis l'env.
#
# La télémétrie ne doit JAMAIS pouvoir empêcher le service de démarrer. Le
# 27/08/2026, `langfuse>=2.51.0` résolvait vers la 4.14.5 alors que le code
# appelait l'API v2 (`langfuse.decorators`, supprimée en v3) : l'import au niveau
# module aurait suffi à tuer FastAPI au démarrage, pas à dégrader l'observabilité.
# Deux protections, complémentaires :
#   - `langfuse==4.14.5` épinglé exactement dans requirements.txt (un `>=` sur une
#     bibliothèque qui a cassé son API deux fois rend le build non reproductible) ;
#   - ce garde : import raté, version incompatible ou clés absentes → `observe`
#     devient transparent et le client muet, le backend démarre normalement.
try:
    from langfuse import Langfuse, observe  # noqa: F401
    _langfuse = Langfuse()
    LANGFUSE_ENABLED = True
except Exception as _lf_import_err:  # pragma: no cover — chemin de secours
    print(f"Langfuse indisponible, observabilité désactivée : {_lf_import_err}")
    LANGFUSE_ENABLED = False

    def observe(func=None, **_kwargs):
        """Décorateur transparent — remplace langfuse.observe quand le SDK manque."""
        return func if func is not None else (lambda f: f)

    class _NullLangfuse:
        """Absorbe tout appel, y compris chaîné (start_observation().update().end())."""
        def __getattr__(self, _name):
            def _noop(*_args, **_kwargs):
                return self
            return _noop

    _langfuse = _NullLangfuse()

# Patterns de noms de fichiers indiquant une source administrative/financière.
# Évalués UNE FOIS à l'indexation (index_source) pour écrire is_administrative=True en DB.
# Le RAG filtre ensuite via la colonne — pas de pattern-matching au runtime.
_ADMIN_PATTERNS = [
    "facture", "devis", "bon de commande", "invoice",
    "order", "avoir", "nda", "commande", "proforma",
]


def _is_administrative(source_name: str) -> bool:
    low = source_name.lower()
    return any(p in low for p in _ADMIN_PATTERNS)


def _load_biencoder() -> SentenceTransformer:
    return SentenceTransformer("paraphrase-multilingual-mpnet-base-v2", device="cpu")


def _load_crossencoder() -> CrossEncoder:
    return CrossEncoder("cross-encoder/mmarco-mMiniLMv2-L12-H384-v1", device="cpu")


@asynccontextmanager
async def lifespan(app: FastAPI):
    global model, reranker
    loop = asyncio.get_running_loop()
    print("Loading models (bi-encoder + cross-encoder reranker)...")
    model, reranker = await asyncio.gather(
        loop.run_in_executor(None, _load_biencoder),
        loop.run_in_executor(None, _load_crossencoder),
    )
    print(f"All models loaded. Embedding dim: {len(model.encode('test'))}")
    yield
    # Flush Langfuse au shutdown pour ne pas perdre les dernières traces
    _langfuse.flush()


app = FastAPI(lifespan=lifespan)

ALLOWED_ORIGIN = "https://khadija-benayed.github.io"

app.add_middleware(
    CORSMiddleware,
    allow_origins=[ALLOWED_ORIGIN],
    allow_methods=["POST", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)

_CORS_HEADERS = {
    "Access-Control-Allow-Origin": ALLOWED_ORIGIN,
    "Access-Control-Allow-Headers": "Content-Type, Authorization",
}


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    if request.method == "OPTIONS" or request.url.path == "/health":
        return await call_next(request)

    ip = request.client.host if request.client else "unknown"

    # Authenticate first so the rate-limit bucket is per-user, not per-IP.
    # Users behind the same corporate NAT would otherwise share a bucket.
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:]
        try:
            user_resp = sb.auth.get_user(token)
            user_id = user_resp.user.id if user_resp.user else None
            if not user_id:
                return JSONResponse({"error": "unauthorized"}, status_code=401, headers=_CORS_HEADERS)
        except Exception:
            return JSONResponse({"error": "unauthorized"}, status_code=401, headers=_CORS_HEADERS)
    else:
        return JSONResponse({"error": "unauthorized"}, status_code=401, headers=_CORS_HEADERS)

    if not _check_rate_limit(user_id or ip):
        return JSONResponse({"error": "rate limit exceeded"}, status_code=429, headers=_CORS_HEADERS)

    request.state.user_id = user_id
    try:
        return await call_next(request)
    except Exception as e:
        print(f"Unhandled exception: {e}\n{traceback.format_exc()}")
        return JSONResponse({"error": "internal server error"}, status_code=500, headers=_CORS_HEADERS)

# ── Environment variables ─────────────────────────────────────────────────────
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
if not GOOGLE_API_KEY:
    print("WARNING: GOOGLE_API_KEY not set — Gemini calls will fail")
GOOGLE_SA_KEY = os.environ.get("GOOGLE_SA_KEY")  # JSON string

sb: Client = _make_sb_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
genai.configure(api_key=GOOGLE_API_KEY)

# Désactive tous les filtres de sécurité — app B2B interne, contenu métier légitime
# Choix assumé : désactive tous les filtres Gemini.
# Justification : outil B2B interne (agence cosmétique/bien-être), accès restreint par JWT Supabase,
# contenu métier légitime bloqué à tort par les filtres par défaut (finish_reason=SAFETY).
# À réévaluer si l'app devient accessible au public.
_SAFETY_OFF = {
    HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_NONE,
    HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_NONE,
    HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_NONE,
    HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_NONE,
}

# ── Gemini model IDs ──────────────────────────────────────────────────────────
GEMINI_FLASH = "gemini-2.5-flash"
GEMINI_PRO = "gemini-2.5-pro"

# MMR dedup — cosine threshold above which a chunk is considered redundant with an
# already-selected one.
#
# Calibration en attente (juin 2026) — procédure :
#   1. python eval/run_eval.py --judge --mmr-threshold 1.01   # MMR désactivé
#   2. python eval/run_eval.py --judge                         # MMR actif (0.92)
#   Comparer source_recall. Si (1) >= (2) → MMR dégrade le rappel → retirer ou monter à 0.95.
#   Si (1) < (2) → MMR améliore la diversité → garder, tester 0.90.
#
# Note : pool = ~20-26 chunks (match_count=20 + safety net ≤3×2). Le diversity cap (2/source)
# gère déjà l'intra-source. À 0.92, MMR filtre surtout des quasi-doublons cross-source
# dont la cosine > 0.92 — rare pour du contenu réellement distinct.
MMR_SIM_THRESHOLD = 0.92

# ── Cost calculation ──────────────────────────────────────────────────────────
GEMINI_PRICING: dict[str, dict[str, float]] = {
    GEMINI_FLASH: {
        "input": 0.15 / 1_000_000,
        "output": 0.60 / 1_000_000,
    },
    GEMINI_PRO: {
        "input": 1.25 / 1_000_000,
        "output": 10.00 / 1_000_000,
    },
}


def calculate_cost(model_id: str, usage: Optional[dict]) -> float:
    if not usage:
        return 0.0
    rates = GEMINI_PRICING.get(model_id, GEMINI_PRICING[GEMINI_FLASH])
    return usage.get("input_tokens", 0) * rates["input"] + usage.get("output_tokens", 0) * rates["output"]


_GEMINI_BLOCKED = {"SAFETY", "RECITATION", "PROHIBITED_CONTENT", "BLOCKLIST", "SPII"}

def _gemini_text(response) -> str:
    """Extrait le texte d'une réponse Gemini. Lève ValueError si bloquée ou sans contenu."""
    if not response.candidates:
        raise ValueError("Gemini n'a retourné aucun candidat (réponse vide)")
    finish = response.candidates[0].finish_reason
    finish_name = finish.name if hasattr(finish, "name") else str(finish)
    if finish_name in _GEMINI_BLOCKED:
        raise ValueError(f"Réponse bloquée par Gemini (finish_reason={finish_name})")
    if finish_name == "MAX_TOKENS":
        print(f"_gemini_text: WARNING réponse tronquée (MAX_TOKENS)")
    try:
        return response.text
    except Exception:
        # MAX_TOKENS with thinking models (Gemini 2.5 Pro): response.text may throw
        # if thinking consumed all tokens. Try extracting text parts directly.
        parts = getattr(response.candidates[0].content, "parts", [])
        text_parts = [p.text for p in parts if hasattr(p, "text") and not getattr(p, "thought", False)]
        if text_parts:
            return "".join(text_parts)
        raise ValueError(f"Gemini n'a retourné aucun texte (finish_reason={finish_name})")


# ── Embedding (local, zero timeout) ──────────────────────────────────────────
def embed_texts(texts: list[str]) -> list[list[float]]:
    if model is None:
        raise RuntimeError("Model not loaded")
    embeddings = model.encode(texts, normalize_embeddings=True)
    return embeddings.tolist()


def _rerank_chunks(query: str, chunks: list[dict]) -> list[dict]:
    """Score (query, « nom [date]\n texte ») pairs with the cross-encoder, sorted desc.
    Falls back to input order if reranker is unavailable.

    Le nom de la source fait partie de la paire, comme le préfixe d'`index_source` fait
    partie de ce qui est embeddé et comme `source_name` fait partie du tsvector depuis
    la migration 20260827c. Sans lui le reranker était le dernier étage aveugle au nom
    du fichier, et c'est ce qui restait cassé après ces deux correctifs :

    « que s'est-il passé lors du dernier point de tracking ? » plaçait bien
    « Notes point suivi tracking 31/07/2026 » au rang 1 du pool, mais avec un logit de
    -4.83 — mesuré le 27/08/2026. Or sur le testset le minimum des cas « answer » est
    -0.96 et la médiane des cas « abstain » -1.23 : le cross-encoder classait donc cette
    question parmi celles auxquelles il faut s'abstenir de répondre, et le seuil
    d'injection la rejetait à juste titre selon lui. La pertinence de ces notes ne vit
    pas dans leur corps — des puces techniques sans date ni intitulé — mais dans leur
    nom, que le reranker ne voyait pas.

    Ne PAS compenser en baissant le seuil d'injection : mesuré, il tombe pile sur la
    frontière answer/abstain, et le descendre à -4 injecterait des documents dans 12 des
    14 cas d'abstention du testset.
    """
    if reranker is None or not chunks:
        return chunks

    def _pair_text(c: dict) -> str:
        name = (c.get("source_name") or "")[:60]
        date_tag = ""
        raw = c.get("drive_modified_at")
        if raw:
            try:
                date_tag = f" [{datetime.fromisoformat(str(raw).replace('Z', '+00:00')).strftime('%d/%m/%Y')}]"
            except Exception:
                pass
        prefix = f"{name}{date_tag}\n" if name else ""
        return prefix + (c.get("chunk_text") or "")

    pairs = [(query, _pair_text(c)) for c in chunks]
    scores = reranker.predict(pairs)
    return sorted(
        [dict(c, rerank_score=float(s)) for c, s in zip(chunks, scores)],
        key=lambda c: c["rerank_score"],
        reverse=True,
    )


def _mmr_filter(chunks: list, embeddings: list, threshold: float) -> list:
    """Discard chunks whose cosine similarity to any already-selected chunk exceeds
    `threshold`. Embeddings must be pre-normalized (dot product == cosine). Chunks
    must be sorted by final_score descending before calling.
    Chunks with None embedding are kept unconditionally (safety net fallback)."""
    selected, selected_embs = [], []
    for chunk, emb in zip(chunks, embeddings):
        if emb is None:
            selected.append(chunk)
            continue
        emb_arr = np.array(emb, dtype="float32")
        if selected_embs and max(float(np.dot(emb_arr, s)) for s in selected_embs) > threshold:
            continue
        selected.append(chunk)
        selected_embs.append(emb_arr)
    return selected


# ── Sources d'un client ───────────────────────────────────────────────────────
def _client_source_names(client_id: str, page_size: int = 1000, max_pages: int = 50) -> set:
    """Noms de sources distincts d'un client, éligibles au RAG.

    PostgREST ne sait pas faire DISTINCT : on pagine sur la seule colonne
    `source_name` et on déduplique en Python.

    Un `.limit(N)` ne convient pas — il tronque les CHUNKS, pas les sources.
    Mesuré le 27/08/2026 sur aroma-zone (5 110 chunks, 123 sources) : le
    `.limit(500)` d'avant ne révélait que 17 sources. Le filet de sécurité par
    nom de fichier, censé rattraper un document nommé dans la question, était
    donc aveugle à 86 % du corpus et ne se déclenchait quasiment jamais.

    `order("source_name")` rend la pagination `range()` déterministe.

    is_administrative exclu : le filet n'a que 3 emplacements, et le fetch par source
    filtre déjà cette colonne. Sans ce filtre, une pièce comptable pouvait consommer
    un emplacement pour ne ramener aucune ligne — 33 des 122 sources d'aroma-zone.
    """
    names: set = set()
    for page in range(max_pages):
        start = page * page_size
        rows = (
            sb.table("document_chunks")
            .select("source_name")
            .eq("client_id", client_id)
            .eq("is_administrative", False)
            .not_.is_("source_name", "null")
            .order("source_name")
            .range(start, start + page_size - 1)
            .execute()
        ).data or []
        names.update(r["source_name"] for r in rows)
        if len(rows) < page_size:
            return names
    print(f"_client_source_names: plafond de {max_pages * page_size} chunks atteint "
          f"pour {client_id} — la liste des sources est peut-être incomplète")
    return names


# ── Supabase retry ───────────────────────────────────────────────────────────
def _sb_insert(table: str, rows: list, max_attempts: int = 3) -> None:
    """Insert rows with up to max_attempts retries on transient network errors."""
    for attempt in range(max_attempts):
        try:
            sb.table(table).insert(rows).execute()
            return
        except Exception:
            if attempt < max_attempts - 1:
                time.sleep(attempt + 1)  # 1s, 2s
            else:
                raise


# ── Rate limiting ─────────────────────────────────────────────────────────────
_rate_lock = threading.Lock()
_rate_buckets: dict = defaultdict(list)
RATE_LIMIT = 60  # requests per minute per IP


def _check_rate_limit(key: str) -> bool:
    now = time.time()
    cutoff = now - 60
    with _rate_lock:
        bucket = [t for t in _rate_buckets.get(key, []) if t > cutoff]
        if not bucket and key in _rate_buckets:
            del _rate_buckets[key]
        if len(bucket) >= RATE_LIMIT:
            _rate_buckets[key] = bucket
            return False
        bucket.append(now)
        _rate_buckets[key] = bucket
        return True


# ── Chunking ──────────────────────────────────────────────────────────────────
def chunk_text(text: str, max_chars: int = 1200, overlap: int = 200) -> list[str]:
    """
    Splits text into chunks preserving sentence boundaries.
    Priority: paragraph breaks → sentence ends → character limit.
    max_chars=1200 ≈ 300 tokens (well within mpnet-base-v2's 512-token limit).
    """
    normalized = re.sub(r"\r\n|\r", "\n", text).strip()
    if not normalized:
        return []

    lines = normalized.split('\n')
    result_lines = []
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('#') and i > 0:
            result_lines.append('')
        elif (stripped.startswith('* ') or stripped.startswith('- ')) and i > 0:
            prev = lines[i - 1].strip()
            if prev and not prev.startswith('* ') and not prev.startswith('- '):
                result_lines.append('')
        result_lines.append(line)
    normalized = '\n'.join(result_lines)

    segments: list[str] = []
    for para in re.split(r"\n{2,}", normalized):
        p = para.strip()
        if not p:
            continue
        if len(p) <= max_chars:
            segments.append(p)
        else:
            buf = ""
            for part in re.split(r"(?<=[.!?])\s+|\n", p):
                part = part.strip()
                if not part:
                    continue
                candidate = (buf + " " + part) if buf else part
                if len(candidate) > max_chars and buf:
                    segments.append(buf.strip())
                    buf = part
                else:
                    buf = candidate
            if buf.strip():
                segments.append(buf.strip())

    chunks: list[str] = []
    current = ""
    for seg in segments:
        joined = (current + "\n\n" + seg) if current else seg
        if len(joined) > max_chars and current:
            chunks.append(current)
            tail = current[-overlap:]
            m = re.search(r"(?<=[.!?])\s", tail)
            overlap_text = tail[m.end() :] if m else tail
            current = (overlap_text + "\n\n" + seg) if overlap_text else seg
        else:
            current = joined

    if current.strip():
        chunks.append(current.strip())

    return chunks if chunks else [normalized[:max_chars]]


def chunk_csv(text: str, max_chars: int = 500) -> list[str]:
    """
    Chunks CSV with header repeated in each chunk.
    Wide sheets (many columns) → fewer rows per chunk; narrow sheets → more rows per chunk.
    """
    lines = [l for l in text.split("\n") if l.strip()]
    if len(lines) <= 1:
        return lines
    header = lines[0][:max_chars - 50]  # truncate pathologically wide headers
    budget = max(50, max_chars - len(header) - 1)
    chunks: list[str] = []
    current_rows: list[str] = []
    current_len = 0
    for line in lines[1:]:
        row = line if len(line) <= budget else line[:budget]
        row_len = len(row) + 1
        if current_rows and current_len + row_len > budget:
            chunks.append(header + "\n" + "\n".join(current_rows))
            current_rows = []
            current_len = 0
        current_rows.append(row)
        current_len += row_len
    if current_rows:
        chunks.append(header + "\n" + "\n".join(current_rows))
    return chunks


# ── Google Drive ──────────────────────────────────────────────────────────────
def get_drive_service() -> tuple:
    """Returns (drive_service, sa_email). Raises ValueError if GOOGLE_SA_KEY missing."""
    if not GOOGLE_SA_KEY:
        raise ValueError("GOOGLE_SA_KEY manquante")
    sa_info = json.loads(GOOGLE_SA_KEY)
    creds = service_account.Credentials.from_service_account_info(
        sa_info,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    return drive, sa_info.get("client_email", "")


def _fix_cell(s: str) -> str:
    """Corrige le double-encodage UTF-8→Latin-1 fréquent dans les sheets issus de CSV Windows.
    Ex: 'Ã©' → 'é'. Non-destructif : retourne s intact si la correction échoue."""
    try:
        return s.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s


def _export_large_sheet(file_id: str, file_name: str) -> Optional[str]:
    """Fallback for Google Sheets that exceed the Drive API 10 MB export limit.
    Lists all visible tabs via the Sheets API, exports each as CSV via authenticated
    HTTP, and prefixes every block with [Fichier] + [Onglet] for RAG context.
    """
    from google.auth.transport.requests import AuthorizedSession

    if not GOOGLE_SA_KEY:
        return None
    sa_info = json.loads(GOOGLE_SA_KEY)
    creds = service_account.Credentials.from_service_account_info(
        sa_info,
        scopes=[
            "https://www.googleapis.com/auth/drive.readonly",
            "https://www.googleapis.com/auth/spreadsheets.readonly",
        ],
    )
    sheets_svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
    meta = sheets_svc.spreadsheets().get(
        spreadsheetId=file_id,
        fields="sheets.properties(sheetId,title,hidden)",
    ).execute()

    tabs = [
        (s["properties"]["sheetId"], s["properties"]["title"])
        for s in meta.get("sheets", [])
        if not s["properties"].get("hidden", False)
    ]
    if not tabs:
        return None

    session = AuthorizedSession(creds)
    parts = []
    for gid, title in tabs:
        url = (
            f"https://docs.google.com/spreadsheets/d/{file_id}"
            f"/export?format=csv&gid={gid}"
        )
        resp = session.get(url, timeout=30)
        if resp.status_code != 200:
            print(f"  CSV tab '{title}' → HTTP {resp.status_code}, skipped")
            continue
        csv_text = "\n".join(_fix_cell(line) for line in resp.text.strip().splitlines())
        if not csv_text:
            continue
        parts.append(f"[Fichier : {file_name}] [Onglet : {title}]\n{csv_text}")

    return ("\n\n".join(parts))[:100_000] if parts else None


# ── Gmail (domain-wide delegation) ───────────────────────────────────────────
def get_gmail_service(user_email: str):
    """Gmail API service impersonating a Google Workspace user. Requires DWD on SA."""
    if not GOOGLE_SA_KEY:
        raise ValueError("GOOGLE_SA_KEY manquante")
    sa_info = json.loads(GOOGLE_SA_KEY)
    creds = service_account.Credentials.from_service_account_info(
        sa_info,
        scopes=["https://www.googleapis.com/auth/gmail.readonly"],
    ).with_subject(user_email)
    return build("gmail", "v1", credentials=creds, cache_discovery=False)


def _extract_email_text(payload: dict) -> str:
    """Extracts plain text from a Gmail message payload. Prefers text/plain, strips HTML fallback."""
    import html as _html

    def _walk(part: dict) -> str:
        mime = part.get("mimeType", "")
        if mime == "text/plain":
            data = part.get("body", {}).get("data", "")
            if data:
                return base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
        if mime == "text/html":
            data = part.get("body", {}).get("data", "")
            if data:
                raw = base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
                return _html.unescape(re.sub(r"<[^>]+>", " ", raw))
        for sub in part.get("parts", []):
            result = _walk(sub)
            if result.strip():
                return result
        return ""

    text = _walk(payload)
    return re.sub(r"\s+", " ", text).strip()[:2500]


_OFFICE_MIME = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/msword": "docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xlsx",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": "pptx",
    "text/plain": "txt",
    "text/csv": "csv",
}

# Sentinel retourné par export_drive_file quand une exception imprévue survient
# (réseau, fichier corrompu…). Distinct de None qui signifie "MIME non supporté"
# (skip permanent). Le caller traite _EXPORT_TRANSIENT_ERR comme une erreur
# transitoire : compteur errors++, pas d'entrée sync_ignored → le fichier sera
# retrigué au prochain sync.
_EXPORT_TRANSIENT_ERR = object()

# Raisons écrites dans sync_ignored qui représentent un échec définitif.
# Ces fichiers sont traités comme "cached" dans _is_cached pour éviter de
# gaspiller 150 s de slot à chaque sync sur un fichier qui ne peut jamais
# être indexé. Les raisons transitoires ('timeout', 'error') ne sont PAS
# dans cet ensemble : elles restent retriables.
_PERM_IGNORE_REASONS = frozenset({"empty", "skipped", "ineligible_ai", "export_error", "ineligible"})


def export_drive_file(file_id: str, file_name: str, mime_type: str) -> Optional[dict]:
    """
    Downloads and extracts a Drive file to plain text.
    Creates its own Drive service — safe for concurrent calls from multiple threads.
    Returns None for unsupported types (images, videos, etc.).
    """
    try:
        drive, _ = get_drive_service()

        # Google Workspace → export as text
        if mime_type == "application/vnd.google-apps.spreadsheet":
            import io
            import openpyxl
            try:
                req = drive.files().export_media(
                    fileId=file_id,
                    mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                raw = _download_bytes(req)
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        cells = [_fix_cell(str(c)) if c is not None else "" for c in row]
                        if any(cells):
                            rows.append("\t".join(cells))
                    if rows:
                        parts.append(
                            f"[Fichier : {file_name}] [Onglet : {sheet_name}]\n"
                            + "\n".join(rows)
                        )
                wb.close()
                content = "\n\n".join(parts)[:100_000]
                if not content.strip():
                    print(f"Sheet export empty for {file_name}")
                return {"filename": file_name, "type": "csv", "content": content}
            except _GHttpError as http_err:
                reason = (
                    http_err.error_details[0].get("reason", "")
                    if http_err.error_details
                    else str(http_err)
                )
                if reason in ("exportSizeLimitExceeded", "cannotExportFile"):
                    print(
                        f"Sheet too large for xlsx export ({reason}), "
                        f"falling back to per-tab CSV: {file_name}"
                    )
                    content = _export_large_sheet(file_id, file_name)
                    if content:
                        return {"filename": file_name, "type": "csv", "content": content}
                    print(f"Per-tab CSV fallback also failed for {file_name}")
                else:
                    print(f"Sheet export HttpError for {file_name}: {http_err}")
                return None

        if mime_type in ("application/vnd.google-apps.document", "application/vnd.google-apps.presentation"):
            req = drive.files().export_media(fileId=file_id, mimeType="text/plain")
            content = _download_bytes(req).decode("utf-8", errors="replace")[:20_000]
            return {"filename": file_name, "type": "txt", "content": content}

        # PDF → subprocess-isolated extraction
        if mime_type == "application/pdf":
            raw = _download_bytes(drive.files().get_media(fileId=file_id))
            content = safe_extract(raw, mime_type)
            if not content.strip():
                return None
            return {"filename": file_name, "type": "pdf", "content": content[:50_000]}

        # Office / plain text → subprocess-isolated extraction
        if mime_type in _OFFICE_MIME:
            ext = _OFFICE_MIME[mime_type]
            raw = _download_bytes(drive.files().get_media(fileId=file_id))
            content = safe_extract(raw, mime_type)
            if not content.strip():
                return None
            type_label = {"docx": "doc", "xlsx": "sheet", "pptx": "ppt"}.get(ext, ext)
            cap = 100_000 if type_label == "sheet" else 20_000
            return {"filename": file_name, "type": type_label, "content": content[:cap]}

        return None

    except Exception as e:
        print(f"Export error for {file_name}: {e}")
        return _EXPORT_TRANSIENT_ERR


def _list_files_recursive(
    drive, folder_id: str, visited: set, max_files: int, total_ref: list
) -> list:
    """Recursively lists all non-folder files in a Drive folder with pagination."""
    if folder_id in visited or total_ref[0] >= max_files:
        return []
    visited.add(folder_id)

    all_items = []
    page_token = None
    while True:
        params = {
            "q": f"'{folder_id}' in parents and trashed = false",
            "fields": "nextPageToken,files(id,name,mimeType,modifiedTime)",
            "pageSize": 100,
            "supportsAllDrives": True,
            "includeItemsFromAllDrives": True,
        }
        if page_token:
            params["pageToken"] = page_token
        result = drive.files().list(**params).execute()
        all_items.extend(result.get("files", []))
        page_token = result.get("nextPageToken")
        if not page_token:
            break

    files = [f for f in all_items if f["mimeType"] != "application/vnd.google-apps.folder"]
    sub_folders = [f for f in all_items if f["mimeType"] == "application/vnd.google-apps.folder"]
    total_ref[0] += len(files)

    sub_files = []
    for sf in sub_folders:
        if total_ref[0] >= max_files:
            break
        sub_files.extend(_list_files_recursive(drive, sf["id"], visited, max_files, total_ref))

    return files + sub_files


def _type_priority(mime_type: str) -> int:
    if mime_type in (
        "application/vnd.google-apps.spreadsheet",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/csv",
    ):
        return 0
    if mime_type in (
        "application/vnd.google-apps.document",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
        "text/plain",
    ):
        return 1
    if mime_type in (
        "application/vnd.google-apps.presentation",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ):
        return 2
    if mime_type == "application/pdf":
        return 3
    return 4


def _parse_modified(f: dict) -> datetime:
    try:
        return datetime.fromisoformat(f.get("modifiedTime", "").replace("Z", "+00:00"))
    except Exception:
        return datetime.min.replace(tzinfo=timezone.utc)


# Date portée par le nom du fichier. Deux formats, année sur 4 chiffres seulement.
# La DERNIÈRE occurrence est retenue : sur une plage (« Audit 10/10/23 - 10/04/24 »)
# la date de fin est la plus parlante. Les années sur 2 chiffres sont ignorées —
# « 15_05_24 » est ambigu, mieux vaut retomber sur drive_modified_at que deviner.
# Doit rester aligné sur le backfill SQL de la migration 20260827d.
_DOC_DATE_DMY = re.compile(r".*([0-3]\d)/([01]\d)/(20\d{2})")
_DOC_DATE_YMD = re.compile(r".*(20\d{2})-([01]\d)-([0-3]\d)")


def _doc_date_from_name(source_name: str | None) -> Optional[str]:
    """ISO 8601 (midi UTC) si le nom porte une date, sinon None.

    Pourquoi : `modifiedTime` de Drive dit quand le fichier a été touché, pas de
    quand date son contenu. Les 6 « Notes point suivi tracking » d'aroma-zone ont
    toutes drive_modified_at = 2026-07-31 (rangées le même jour) alors qu'elles
    couvrent avril à juillet. Sans cette extraction, « le dernier point » n'a
    aucun signal sur lequel s'appuyer.

    Midi UTC volontaire : évite qu'un décalage de fuseau fasse basculer la date
    d'un jour dans un sens ou dans l'autre.
    """
    if not source_name:
        return None
    m = _DOC_DATE_DMY.match(source_name)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
    else:
        m = _DOC_DATE_YMD.match(source_name)
        if not m:
            return None
        y, mo, d = m.group(1), m.group(2), m.group(3)
    try:
        return datetime(int(y), int(mo), int(d), 12, 0, 0, tzinfo=timezone.utc).isoformat()
    except ValueError:
        return None  # 31/02/2026 et autres dates impossibles


def _temporal_score(drive_modified_at: str | None, decay_days: int = 180) -> float:
    if not drive_modified_at:
        return 0.5
    dt = _parse_modified({"modifiedTime": drive_modified_at})
    if dt == datetime.min.replace(tzinfo=timezone.utc):
        return 0.5
    age_days = (datetime.now(timezone.utc) - dt).days
    return math.exp(-age_days / decay_days)


# ── Trace de diagnostic RAG ───────────────────────────────────────────────────
# Désactivée par défaut, activée par RAG_DEBUG=1 sur le service Cloud Run.
#
# Le pipeline de recherche n'avait aucune observabilité : quatre print() sur 400
# lignes, tous sur des chemins d'erreur. Diagnostiquer « pourquoi ce document ne
# remonte pas » se faisait donc en relisant le code et en supposant. Cette trace
# montre, requête par requête, ce qui est récupéré, comment c'est classé, et ce
# qui est finalement injecté dans le prompt.
#
# flush=True volontaire : sans lui, stdout est bufferisé par blocs et les lignes
# ressortent groupées dans Cloud Logging, avec des horodatages inexploitables.
RAG_DEBUG = os.getenv("RAG_DEBUG") == "1"


def _rag_log(line: str) -> None:
    if RAG_DEBUG:
        print(f"[RAG] {line}", flush=True)


def _rag_log_pool(title: str, rows: list, limit: int = 15) -> None:
    """Tableau compact : scores, type, date, source. Ne doit jamais lever."""
    if not RAG_DEBUG:
        return
    try:
        _rag_log(f"{title} — {len(rows)} chunks")
        _rag_log(f"  {'rerank':>8} {'temp':>6} {'final':>8}  {'type':<14} {'date':<10} source")
        _nan = float("nan")
        for c in rows[:limit]:
            _d = str(c.get("drive_modified_at") or "")[:10] or "—"
            _rag_log(
                f"  {c.get('rerank_score', _nan):>8.3f}"
                f" {c.get('_temporal', _nan):>6.2f}"
                f" {c.get('final_score', _nan):>8.3f}"
                f"  {str(c.get('source_type'))[:14]:<14} {_d:<10} {str(c.get('source_name'))[:60]}"
            )
        if len(rows) > limit:
            _rag_log(f"  … et {len(rows) - limit} autres")
    except Exception as exc:
        print(f"[RAG] trace error (non bloquant): {exc}", flush=True)


# ── Local file extraction ──────────────────────────────────────────────────────
def _download_bytes(req) -> bytes:
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


_WORKER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "extract_worker.py")


def safe_extract(file_bytes: bytes, mime_type: str) -> str:
    """Run file extraction in an isolated subprocess.

    Any native crash (SIGABRT, SIGSEGV) in the child does NOT kill the server.
    Returns empty string on crash, timeout, or unsupported type.
    """
    try:
        result = subprocess.run(
            [sys.executable, _WORKER, mime_type],
            input=base64.b64encode(file_bytes),
            capture_output=True,
            timeout=120,
        )
        return result.stdout.decode("utf-8", errors="replace") if result.returncode == 0 else ""
    except subprocess.TimeoutExpired:
        print(f"safe_extract timeout for {mime_type}")
        return ""
    except Exception as e:
        print(f"safe_extract error: {e}")
        return ""


# ── eval_judge ────────────────────────────────────────────────────────────────
_EVAL_JUDGE_ANSWER_SCHEMA = {
    "type": "object",
    "properties": {
        "correctness": {"type": "number"},
        "faithful":    {"type": "number"},
        "reasoning":   {"type": "string"},
    },
    "required": ["correctness", "faithful", "reasoning"],
}

_EVAL_JUDGE_ABSTAIN_SCHEMA = {
    "type": "object",
    "properties": {
        "abstained_properly": {"type": "number"},
        "fabricated":         {"type": "number"},
        "reasoning":          {"type": "string"},
    },
    "required": ["abstained_properly", "fabricated", "reasoning"],
}

@observe(name="eval_judge", as_type="generation")
async def eval_judge(body: dict) -> JSONResponse:
    question = body.get("question", "")
    context = body.get("context", "")
    response = body.get("response", "")
    expected_points = body.get("expected_points", [])
    mode = body.get("mode", "answer")

    if not question or not response:
        return JSONResponse({"error": "question et response sont requis"}, status_code=400)

    if mode == "answer":
        if not context:
            context = "(aucun document injecté — évaluer uniquement sur la cohérence de la réponse)"
        prompt = (
            f"Tu es un juge d'évaluation RAG. Évalue la réponse ci-dessous.\n\n"
            f"Question : {question}\n\n"
            f"Contexte injecté :\n{context}\n\n"
            f"Réponse du modèle :\n{response}\n\n"
            f"Points attendus : {json.dumps(expected_points, ensure_ascii=False)}\n\n"
            "Score `correctness` (0.0–1.0) : proportion des points attendus couverts sémantiquement "
            "(une reformulation correcte compte comme couvert).\n"
            "Score `faithful` (0.0–1.0) : la réponse ne contredit pas le contexte injecté "
            "(1.0 = aucune contradiction, 0.0 = contradiction directe).\n"
            "Retourne uniquement le JSON structuré demandé. "
            "Le champ reasoning doit rester concis : 3-4 phrases maximum, pas de citation verbatim du contexte."
        )
        schema = _EVAL_JUDGE_ANSWER_SCHEMA
    else:  # abstain
        prompt = (
            f"Tu es un juge d'évaluation RAG. Évalue si le modèle s'est correctement abstenu.\n\n"
            f"Question : {question}\n\n"
            f"Contexte injecté (vide = aucun doc pertinent) :\n{context}\n\n"
            f"Réponse du modèle :\n{response}\n\n"
            "Score `abstained_properly` (0.0–1.0) : le modèle a bien signalé l'absence d'information "
            "sans inventer (1.0 = abstention claire et correcte).\n"
            "Score `fabricated` (0.0–1.0) : proportion d'informations inventées dans la réponse "
            "(0.0 = rien d'inventé, 1.0 = tout est inventé — score inversé, plus bas = mieux).\n"
            "Retourne uniquement le JSON structuré demandé. "
            "Le champ reasoning doit rester concis : 3-4 phrases maximum, pas de citation verbatim du contexte."
        )
        schema = _EVAL_JUDGE_ABSTAIN_SCHEMA

    try:
        gemini = genai.GenerativeModel(
            model_name=GEMINI_FLASH,
            generation_config={
                "temperature": 0,
                "max_output_tokens": 4096,
                "response_mime_type": "application/json",
                "response_schema": schema,
            },
            safety_settings=_SAFETY_OFF,
        )
        resp = gemini.generate_content(prompt)
        finish = resp.candidates[0].finish_reason
        finish_name = finish.name if hasattr(finish, "name") else str(finish)
        if finish_name == "MAX_TOKENS":
            return JSONResponse({"error": "eval_judge error: MAX_TOKENS — réponse JSON tronquée"}, status_code=502)
        result = json.loads(_gemini_text(resp))
    except Exception as e:
        return JSONResponse({"error": f"eval_judge error: {e}"}, status_code=502)

    # ── Langfuse (non bloquant) ────────────────────────────────────────────
    try:
        _um = getattr(resp, "usage_metadata", None)
        _langfuse.update_current_generation(
            input=prompt[:3000],
            output=json.dumps(result)[:2000],
            model=GEMINI_FLASH,
            usage_details={"input": getattr(_um, "prompt_token_count", 0) if _um else 0,
                   "output": getattr(_um, "candidates_token_count", 0) if _um else 0},
            metadata={"case_id": body.get("case_id", "unknown")},
        )
    except Exception:
        pass

    return JSONResponse(result)


# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    # langfuse_enabled : le garde d'import est fail-open et donc silencieux. Sans
    # cette clé, un retour au mode dégradé (SDK absent ou incompatible) ne se voit
    # nulle part. À vérifier après chaque déploiement qui touche la dépendance.
    return {"ok": True, "model_loaded": model is not None,
            "reranker_loaded": reranker is not None,
            "langfuse_enabled": LANGFUSE_ENABLED}


@app.post("/")
async def dispatcher(request: Request):
    body = await request.json()
    action = body.get("action")
    user_id = getattr(request.state, "user_id", None)

    if action == "me":
        return await get_me(request)
    if action == "summarize_session":
        return await summarize_session(body)
    if action == "list_drive_metadata":
        return await list_drive_metadata(body)
    if action == "generate_brief":
        return await generate_brief(body, user_id)
    if action == "index_source":
        return await index_source(body, user_id)
    if action == "delete_source_chunks":
        return await delete_source_chunks(body)
    if action == "save_to_kb":
        return await save_to_kb(body)
    if action == "get_client_members":
        return await get_client_members(body, user_id)
    if action == "add_client_member":
        return await add_client_member(body, user_id)
    if action == "remove_client_member":
        return await remove_client_member(body, user_id)
    if action == "set_member_role":
        return await set_member_role(body, user_id)
    if action == "claim_ownership":
        return await claim_ownership(body, user_id)
    if action == "upsert_task":
        return await upsert_task(body, user_id)
    if action == "propose_cr_tasks":
        return await propose_cr_tasks(body, user_id)
    if action == "delete_task":
        return await delete_task(body, user_id)
    if action == "weekly_digest":
        return await weekly_digest(body, user_id)
    if action == "create_client":
        return await create_client(body, user_id)
    if action == "delete_client":
        return await delete_client(body, user_id)
    if action == "create_invitation":
        return await create_invitation(body, user_id)
    if action == "join_client_via_token":
        return await join_client_via_token(body, user_id)
    if action == "sync_drive":
        return await sync_drive(body, request)
    if action == "sync_state":
        key = f"{body.get('client_id')}|{body.get('folder_id')}"
        return _sync_state.get(key) or JSONResponse({"error": "aucun sync en cours ou récent"}, status_code=404)
    if action == "sync_emails":
        return await sync_emails(body, request)
    if action == "update_gmail_sync":
        return await update_gmail_sync(body, user_id)
    if action == "eval_judge":
        return await eval_judge(body)
    if action is None:
        return await chat(body, user_id=user_id)
    return JSONResponse({"error": f"Action inconnue : {action}"}, status_code=400, headers=_CORS_HEADERS)


# ── /me — current user info + assigned clients ────────────────────────────
async def get_me(request: Request):
    user_id = getattr(request.state, "user_id", None)
    if not user_id:
        return JSONResponse({"error": "JWT requis pour /me"}, status_code=401)

    try:
        member_resp = sb.table("team_members").select("*").eq("id", user_id).maybe_single().execute()
        member = member_resp.data

        clients_resp = (
            sb.table("client_members")
            .select("role, clients(id, name, drive_folder_id, context, sources, members)")
            .eq("member_id", user_id)
            .execute()
        )
        clients = []
        for row in (clients_resp.data or []):
            client_data = row.get("clients")
            if client_data:
                clients.append({**client_data, "role": row["role"]})

        return {"member": member, "clients": clients}
    except Exception as e:
        print(f"get_me error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


# ── summarize_session ─────────────────────────────────────────────────────────
async def summarize_session(body: dict):
    client_id = body.get("client_id")
    history = body.get("history", [])

    if not client_id or not history:
        return JSONResponse({"error": "client_id et history requis"}, status_code=400)

    history_text = "\n".join(
        f"{'Utilisateur' if m['role'] == 'u' else 'Assistant'} : {m['text']}"
        for m in history
    )

    try:
        gemini = genai.GenerativeModel(
            model_name=GEMINI_FLASH,
            system_instruction=(
                "Tu es un assistant qui résume des sessions de travail de manière factuelle et concise. "
                "Tu reçois un historique de conversation et tu produis un résumé structuré."
            ),
            generation_config={"max_output_tokens": 600},
            safety_settings=_SAFETY_OFF,
        )
        response = gemini.generate_content(
            "Résume cette session en 5 points max : décisions prises, infos importantes, "
            "actions à faire. Format : liste à tirets, sois factuel et concis. Ne mets pas de titre.\n\n"
            "Session :\n" + history_text
        )
        summary_text = _gemini_text(response)
    except Exception as e:
        return JSONResponse({"error": f"Erreur IA (résumé) : {e}"}, status_code=502)

    try:
        sb.table("session_summaries").insert({"client_id": client_id, "summary_text": summary_text}).execute()
    except Exception as e:
        return JSONResponse({"saved": False, "summary": summary_text, "error": str(e)})

    # CC-208 — Index summary in document_chunks (source_type='session') for semantic search
    try:
        loop = asyncio.get_running_loop()
        embedding = (await loop.run_in_executor(_EMBED_EXECUTOR, embed_texts, [summary_text]))[0]
        session_source_name = f"Session du {time.strftime('%Y-%m-%d')}"
        session_source_id = f"session-{client_id}"
        # Upsert ancré sur source_id — évite d'écraser plusieurs lignes session
        # si des doublons existent (ex : race condition entre deux appels simultanés).
        _updated = sb.table("document_chunks")\
            .update({"source_name": session_source_name, "chunk_text": summary_text, "embedding": embedding})\
            .eq("client_id", client_id).eq("source_type", "session").eq("source_id", session_source_id)\
            .execute()
        if not _updated.data:
            sb.table("document_chunks").insert({
                "client_id": client_id,
                "source_type": "session",
                "source_id": session_source_id,
                "source_name": session_source_name,
                "chunk_text": summary_text,
                "embedding": embedding,
            }).execute()
    except Exception as e:
        print(f"CC-208 index session error (non bloquant): {e}")

    return {"saved": True, "summary": summary_text}


# ── list_drive_metadata ───────────────────────────────────────────────────────
async def list_drive_metadata(body: dict):
    """Lightweight metadata-only listing — no file content download."""
    folder_id = body.get("folder_id")
    if not folder_id:
        return JSONResponse({"error": "folder_id requis"}, status_code=400)

    try:
        drive, sa_email = get_drive_service()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    total_ref = [0]
    all_files = _list_files_recursive(drive, folder_id, set(), 500, total_ref)

    meta_files = [
        {
            "id": f["id"],
            "name": f["name"],
            "mimeType": f["mimeType"],
            "modifiedTime": f.get("modifiedTime", ""),
        }
        for f in all_files
    ]
    return {"files": meta_files, "sa_email": sa_email}


# ── generate_brief ────────────────────────────────────────────────────────────
AGENCY_DOMAIN = "smart-bees.fr"
AGENCY_NAME = "Smart Bees"
CR_KEYWORDS = frozenset({"news", "nouvelles", "réunion", "point",
                         "suivi", "compte", "rendu", "notes", "évoqué", "discuté", "abordé"})

# Structured-output schema for Gemini — forces each field to exist; nullable = null when absent from docs
_BRIEF_SCHEMA = {
    "type": "object",
    "properties": {
        "secteur":           {"type": "string"},
        "enjeux_principaux": {"type": "array", "items": {"type": "string"}},
        "kpis":              {"type": "array", "items": {"type": "string"}},
        "equipe": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "prenom":     {"type": "string"},
                    "nom":        {"type": "string",  "nullable": True},
                    "role":       {"type": "string",  "nullable": True},
                    "email":      {"type": "string",  "nullable": True},
                    "scope":      {"type": "string",  "nullable": True, "enum": ["internal", "external", "uncertain"]},
                    "reports_to": {"type": "string",  "nullable": True},
                },
                "required": ["prenom"],
            },
        },
        "historique": {"type": "string"},
        "notes":      {"type": "string"},
    },
    "required": ["secteur", "enjeux_principaux", "kpis", "equipe", "historique", "notes"],
}


# ── propose_cr_tasks ─────────────────────────────────────────────────────────
_CR_PROPOSAL_SCHEMA = {
    "type": "object",
    "properties": {
        "items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "summary":               {"type": "string"},
                    "match_type":            {"type": "string"},
                    "task_id":               {"type": "integer", "nullable": True},
                    "scope":                 {"type": "string"},
                    "fields": {
                        "type": "object",
                        "properties": {
                            "title":    {"type": "string",  "nullable": True},
                            "assignee": {"type": "string",  "nullable": True},
                            "prio":     {"type": "string",  "nullable": True},
                            "status":   {"type": "string",  "nullable": True},
                            "due_date": {"type": "string",  "nullable": True},
                            "note":     {"type": "string",  "nullable": True},
                        },
                        "required": ["title", "assignee", "prio", "status", "due_date", "note"],
                    },
                    "confidence":             {"type": "number"},
                    "needs_clarification":    {"type": "boolean"},
                    "clarification_question": {"type": "string", "nullable": True},
                },
                "required": [
                    "summary", "match_type", "scope", "fields",
                    "confidence", "needs_clarification", "clarification_question",
                ],
            },
        }
    },
    "required": ["items"],
}


@observe(name="generate_brief", as_type="generation")
async def generate_brief(body: dict, user_id: Optional[str] = None):
    client_id = body.get("client_id")
    docs_content = body.get("docs_content", [])
    existing_brief = body.get("existing_brief")

    if not client_id or not docs_content:
        return JSONResponse(
            {"error": "client_id et docs_content (array non vide) requis"}, status_code=400
        )
    await _assert_role(user_id, client_id, ["owner", "member"])

    TOKEN_BUDGET = 120_000
    total_chars = 0
    doc_blocks = []
    for doc in docs_content:
        block = f"### {doc['filename']}"
        if doc.get("modified_at"):
            try:
                dt = datetime.fromisoformat(doc["modified_at"].replace("Z", "+00:00"))
                block += f" (modifié le {dt.strftime('%d/%m/%Y')})"
            except Exception:
                pass
        block += f"\n{doc['content']}"
        if total_chars + len(block) > TOKEN_BUDGET:
            break
        doc_blocks.append(block)
        total_chars += len(block)

    docs_text = "\n\n---\n\n".join(doc_blocks)

    if existing_brief:
        try:
            existing_brief_block = json.dumps(json.loads(existing_brief), ensure_ascii=False, indent=2)
        except Exception:
            existing_brief_block = existing_brief
    else:
        existing_brief_block = None

    brief_prompt = (
        "Tu es un assistant qui génère une fiche client synthétique pour une agence data/marketing.\n\n"
        + (
            "[FICHE EXISTANTE - à enrichir et corriger si des infos sont outdatées]\n"
            + existing_brief_block + "\n\n"
            if existing_brief_block else ""
        )
        + "[DOCUMENTS CLIENT - classés du plus récent au plus ancien]\n"
        + docs_text + "\n\n"
        + "Génère une fiche client avec exactement ces champs :\n"
        "- secteur : secteur d'activité principal\n"
        "- enjeux_principaux (max 6) : enjeux métier et techniques actuels\n"
        "- kpis (max 6) : indicateurs de performance suivis\n"
        "- equipe : UNIQUEMENT les personnes physiques côté client — PAS les membres Smart Bees, "
        "PAS les outils/plateformes/services (Dacker, Segment, GA4, CommerceTools ne sont PAS des personnes). "
        "Pour chaque contact :\n"
        "  • prenom : requis\n"
        "  • nom : null si absent des docs\n"
        "  • role : poste exact tel que mentionné, null si non précisé — ne pas inventer\n"
        "  • email : adresse si disponible, sinon null\n"
        "  • scope : 'internal' (salarié client), 'external' (prestataire/agence partenaire), 'uncertain' si non précisé\n"
        "  • reports_to : prénom/nom de la personne à qui elle reporte si mentionné, sinon null\n"
        "- historique (3-4 phrases) : chronologie de la collaboration avec dates clés\n"
        "- notes (4-6 phrases) : stack technique, projets en cours et à venir, points d'attention\n\n"
        "Règles :\n"
        "- Prioritise les informations des documents les plus récents\n"
        "- Si la fiche existante contient une info absente des documents, conserve-la\n"
        "- Pour l'équipe, inclure TOUS les contacts client mentionnés dans les documents\n"
        "- Si une information est absente des documents, mettre null — ne pas inventer"
    )

    try:
        gemini = genai.GenerativeModel(
            model_name=GEMINI_PRO,
            generation_config={
                "max_output_tokens": 8192,
                "response_mime_type": "application/json",
                "response_schema": _BRIEF_SCHEMA,
            },
            safety_settings=_SAFETY_OFF,
        )
        response = gemini.generate_content(brief_prompt)
        raw_text = _gemini_text(response)
    except Exception as e:
        return JSONResponse({"error": f"Erreur IA (brief) : {e}"}, status_code=502)

    # Log usage — non-blocking (response is guaranteed non-None here)
    try:
        usage_meta = getattr(response, "usage_metadata", None)
        _in = usage_meta.prompt_token_count if usage_meta else 0
        _out = usage_meta.candidates_token_count if usage_meta else 0
        sb.table("usage_logs").insert({
            "client_id": client_id or None,
            "model": GEMINI_PRO,
            "message_type": "generate_brief",
            "tokens_input": _in,
            "tokens_output": _out,
            "cost_usd": calculate_cost(GEMINI_PRO, {"input_tokens": _in, "output_tokens": _out}),
        }).execute()
    except Exception as e:
        print(f"usage_logs insert error (non bloquant): {e}")

    # ── Langfuse (non bloquant) ────────────────────────────────────────────
    try:
        _langfuse.update_current_generation(
            input=brief_prompt[:3000],
            output=raw_text[:2000],
            model=GEMINI_PRO,
            usage_details={"input": _in, "output": _out},
            metadata={"client_id": client_id,
                       "cost_usd": calculate_cost(GEMINI_PRO, {"input_tokens": _in, "output_tokens": _out})},
        )
    except Exception:
        pass

    try:
        brief = json.loads(raw_text)
    except Exception:
        print(f"generate_brief: JSON invalide reçu de Gemini : {raw_text[:200]}")
        return JSONResponse(
            {"error": "Génération échouée — Gemini n'a pas retourné un JSON valide. Réessaie."},
            status_code=422,
        )

    # Filtre equipe : retire membres agence et faux-positifs outils — conserve les objets
    if "equipe" in brief and isinstance(brief["equipe"], list):
        _TOOL_KW = frozenset({'reporting', 'dashboard', 'analytics', 'tracking',
                              'segment', 'ga4', 'gtm', 'adjust', 'klaviyo', 'dacker'})
        _agency  = AGENCY_NAME.lower()
        def _member_text(m):
            return " ".join(filter(None, [m.get("prenom"), m.get("nom"), m.get("role"), m.get("email")])).lower()
        brief["equipe"] = [
            m for m in brief["equipe"]
            if isinstance(m, dict) and (m.get("prenom") or "").strip()
            and AGENCY_DOMAIN not in _member_text(m)
            and _agency not in _member_text(m)
            and not any(kw in _member_text(m) for kw in _TOOL_KW)
        ]

    expected_keys = ["secteur", "enjeux_principaux", "kpis", "equipe", "historique", "notes"]
    missing = [k for k in expected_keys if k not in brief]
    if missing:
        return JSONResponse(
            {"error": f"Génération incomplète — champs manquants : {', '.join(missing)}. Réessaie."},
            status_code=422,
        )

    try:
        sb.table("clients").update({"context": json.dumps(brief)}).eq("id", client_id).execute()
    except Exception as e:
        return {"brief": brief, "saved": False, "error": str(e)}

    return {"brief": brief, "saved": True}


# ── index_source ──────────────────────────────────────────────────────────────
async def index_source(body: dict, user_id: Optional[str] = None):
    """
    Chunks + embeds (local) + upserts document chunks.
    start_chunk > 0 is a no-op: local embedding processes all chunks in one batch.
    """
    client_id = body.get("client_id")
    if client_id:
        await _assert_role(user_id, client_id, ["owner", "member"])
    source_type = body.get("source_type")
    source_name = body.get("source_name")
    source_id = body.get("source_id")
    drive_modified_at = body.get("drive_modified_at")
    content = body.get("content", "")

    if not source_name or not source_type or not content:
        return JSONResponse(
            {"error": "Paramètres requis : source_type, source_name, content."}, status_code=400
        )
    if not isinstance(content, str) or not content.strip():
        return JSONResponse({"error": "content est vide ou invalide."}, status_code=400)

    # If start_chunk > 0, everything was already processed in the first call
    start_chunk = body.get("start_chunk") or 0
    if start_chunk > 0:
        return {"chunks_created": 0, "has_more": False, "total_chunks": 0}

    text_content = content

    # PDF base64 → subprocess-isolated extraction
    if text_content.startswith("__PDF_BASE64__"):
        pdf_b64 = text_content[len("__PDF_BASE64__"):]
        try:
            pdf_bytes_data = base64.b64decode(pdf_b64)
            loop = asyncio.get_running_loop()
            text_content = await loop.run_in_executor(
                None, safe_extract, pdf_bytes_data, "application/pdf"
            )
            if not text_content.strip():
                raise ValueError("Extraction PDF vide — aucun texte détecté")
        except Exception as e:
            return JSONResponse({"error": f"Extraction PDF échouée : {e}"}, status_code=502)

    is_csv = source_type == "sheet" or (source_name or "").endswith(".csv")
    chunks = chunk_csv(text_content) if is_csv else chunk_text(text_content)

    if not chunks:
        return JSONResponse(
            {"error": "Aucun chunk généré — contenu trop court ou vide."}, status_code=400
        )

    # Embed all at once — local model, zero timeout risk.
    # Prefix = "filename [dd/mm/yyyy]\n" so semantic search benefits from both
    # the document name and its modification date (helps "récent", "dernière version"…).
    loop = asyncio.get_running_loop()
    # doc_date prioritaire sur drive_modified_at, ici comme dans le score temporel :
    # la date du nom qualifie le contenu, modifiedTime seulement le fichier.
    doc_date = _doc_date_from_name(source_name)
    date_tag = ""
    for _candidate in (doc_date, drive_modified_at):
        if not _candidate:
            continue
        try:
            _dt = datetime.fromisoformat(_candidate.replace("Z", "+00:00"))
            date_tag = f" [{_dt.strftime('%d/%m/%Y')}]"
            break
        except Exception:
            continue
    prefix = (source_name[:60] + date_tag + "\n") if source_name else ""
    prefixed_chunks = [(prefix + c) if prefix else c for c in chunks]
    embeddings = await loop.run_in_executor(_EMBED_EXECUTOR, embed_texts, prefixed_chunks)

    # Delete old chunks before inserting (source_id stable key for Drive, source_name fallback)
    try:
        del_q = sb.table("document_chunks").delete()
        if client_id:
            del_q = del_q.eq("client_id", client_id)
        else:
            del_q = del_q.is_("client_id", "null")
        if source_id:
            del_q.eq("source_id", source_id).execute()
        else:
            del_q.eq("source_name", source_name).execute()
    except Exception as e:
        print(f"index_source: delete anciens chunks error (non bloquant): {e}")

    now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    rows = [
        {
            "client_id": client_id or None,
            "source_type": source_type,
            "source_name": source_name,
            **({"source_id": source_id} if source_id else {}),
            **({"drive_modified_at": drive_modified_at} if drive_modified_at else {}),
            **({"doc_date": doc_date} if doc_date else {}),
            "chunk_text": chunk,
            "embedding": emb,
            "last_indexed_at": now,
            "is_administrative": _is_administrative(source_name),
        }
        for chunk, emb in zip(chunks, embeddings)
    ]

    try:
        _sb_insert("document_chunks", rows)
    except Exception as e:
        return JSONResponse({"error": f"Erreur insertion chunks : {e}"}, status_code=500)

    # Log in embedding_logs
    try:
        tokens_estimated = sum(math.ceil(len(c) / 4) for c in chunks)
        sb.table("embedding_logs").insert({
            "client_id": client_id or None,
            "source_name": source_name,
            "chunks_count": len(chunks),
            "tokens_estimated": tokens_estimated,
        }).execute()
    except Exception as e:
        print(f"index_source: embedding_logs insert error (non bloquant): {e}")

    return {"chunks_created": len(chunks), "has_more": False, "total_chunks": len(chunks)}


# ── delete_source_chunks ──────────────────────────────────────────────────────
async def delete_source_chunks(body: dict):
    client_id = body.get("client_id")
    source_name = body.get("source_name")
    source_type_filter = body.get("source_type_filter")

    try:
        if source_type_filter and isinstance(source_type_filter, list):
            q = sb.table("document_chunks").delete()
            if client_id:
                q = q.eq("client_id", client_id)
            q.in_("source_type", source_type_filter).execute()
        elif source_name:
            q = sb.table("document_chunks").delete().eq("source_name", source_name)
            if client_id:
                q = q.eq("client_id", client_id)
            q.execute()
        else:
            return JSONResponse(
                {"error": "source_name ou source_type_filter requis."}, status_code=400
            )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    return {"deleted": True}


# ── save_to_kb ────────────────────────────────────────────────────────────────
async def save_to_kb(body: dict):
    title = body.get("title")
    kb_content = body.get("content")
    source_client = body.get("source_client")
    tags = body.get("tags", [])
    saved_by = body.get("saved_by")

    if not title or not kb_content:
        return JSONResponse({"error": "title et content requis."}, status_code=400)

    try:
        sb.table("agency_knowledge").insert({
            "title": title,
            "content": kb_content,
            "source_client": source_client or None,
            "tags": tags or [],
            "saved_by": saved_by or None,
        }).execute()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    return {"saved": True}


# ── client_members helpers ────────────────────────────────────────────────────
def _is_owner(user_id: Optional[str], client_id: str) -> bool:
    """Returns True if user_id is an owner of client_id. Unauthenticated callers (user_id=None) are never owners."""
    if not user_id:
        return False
    r = sb.table("client_members").select("role").eq("client_id", client_id).eq("member_id", user_id).maybe_single().execute()
    return bool(r.data and r.data.get("role") == "owner")


def _count_owners(client_id: str) -> int:
    r = sb.table("client_members").select("id", count="exact").eq("client_id", client_id).eq("role", "owner").execute()
    return r.count or 0


def _append_member_to_client_json(client_id: str, user_id: str) -> None:
    """Appends user to clients.members JSON. Optimistic concurrency (3 retries) sérialise
    les writes concurrents sur le même champ. Champ cache uniquement — la vérité d'appartenance
    est dans client_members ; une race cross-instance se corrige au prochain selectClient()."""
    try:
        tm = sb.table("team_members").select("full_name, email").eq("id", user_id).maybe_single().execute()
        if not tm.data:
            return
        for _attempt in range(3):
            row = sb.table("clients").select("members").eq("id", client_id).maybe_single().execute()
            if not row.data:
                return
            raw = row.data.get("members")
            current = raw if isinstance(raw, list) else (json.loads(raw) if isinstance(raw, str) and raw else [])
            if any(m.get("member_id") == user_id for m in current):
                return
            used = {m.get("initials", "") for m in current}
            parts = (tm.data.get("full_name") or tm.data.get("email") or "?").strip().split()
            base = (parts[0][0] + parts[1][0]).upper() if len(parts) >= 2 else parts[0][:2].upper() if parts else "?"
            ini, n = base, 2
            while ini in used:
                ini = base + str(n); n += 1
            new_members = current + [{"initials": ini, "name": tm.data.get("full_name") or tm.data.get("email") or "?", "member_id": user_id}]
            # UPDATE conditionnel sur la valeur lue — n'écrit que si aucune autre instance
            # n'a modifié members entre-temps (result.data vide = conflit détecté → retry).
            upd = sb.table("clients").update({"members": json.dumps(new_members)}).eq("id", client_id)
            result = (upd.is_("members", "null") if raw is None else upd.eq("members", json.dumps(current))).execute()
            if result.data:
                return
    except Exception:
        pass


async def _assert_role(user_id: Optional[str], client_id: str, allowed_roles: list):
    """Lève HTTP 403 si user_id n'est pas membre de client_id avec l'un des rôles autorisés."""
    if not user_id:
        raise HTTPException(status_code=401, detail="Non authentifié")
    row = sb.table("client_members") \
        .select("role") \
        .eq("member_id", user_id) \
        .eq("client_id", client_id) \
        .maybe_single().execute()
    if not row.data or row.data["role"] not in allowed_roles:
        raise HTTPException(status_code=403, detail="Accès refusé")



def _recent_note_entries(note_text: str, since_date: str) -> list[str]:
    """Retourne les entrées de note datées >= since_date.

    Supporte deux formats :
    - [YYYY-MM-DD] texte...        (format historique)
    - DD/MM HH:MM — texte...       (format courant dans l'app)
    - DD/MM/YYYY HH:MM — texte...  (variante avec année)
    """
    if not note_text:
        return []

    current_year = datetime.now().year
    parts = re.split(
        r'(?=\[\d{4}-\d{2}-\d{2}\]|(?:^|\n)\d{2}/\d{2}(?:/\d{4})?\s+\d{2}:\d{2}\s*—)',
        note_text,
    )

    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue

        # Format [YYYY-MM-DD]
        m1 = re.match(r'\[(\d{4}-\d{2}-\d{2})\]\s*(.*)', p, re.DOTALL)
        if m1 and m1.group(1) >= since_date:
            txt = m1.group(2).strip().replace('\n', ' ')
            if txt:
                out.append(txt[:400])
            continue

        # Format DD/MM HH:MM — ou DD/MM/YYYY HH:MM —
        m2 = re.match(r'(\d{2})/(\d{2})(?:/(\d{4}))?\s+\d{2}:\d{2}\s*—\s*(.*)', p, re.DOTALL)
        if m2:
            day, month = m2.group(1), m2.group(2)
            year = m2.group(3) or str(current_year)
            iso = f"{year}-{month}-{day}"
            if iso >= since_date:
                txt = m2.group(4).strip().replace('\n', ' ')
                if txt:
                    out.append(txt[:400])

    return out


# ── get_client_members ────────────────────────────────────────────────────────
async def get_client_members(body: dict, user_id: Optional[str]):
    client_id = body.get("client_id")
    if not client_id:
        return JSONResponse({"error": "client_id requis"}, status_code=400)

    try:
        rows = (
            sb.table("client_members")
            .select("member_id, role, team_members(id, email, full_name)")
            .eq("client_id", client_id)
            .execute()
        )
        existing_ids: set = set()
        members = []
        for row in (rows.data or []):
            tm = row.get("team_members") or {}
            members.append({
                "member_id": row["member_id"],
                "role": row["role"],
                "email": tm.get("email", ""),
                "full_name": tm.get("full_name", ""),
            })
            existing_ids.add(row["member_id"])

        all_tm = sb.table("team_members").select("id, email, full_name").execute()
        available = [
            {"id": m["id"], "email": m.get("email", ""), "full_name": m.get("full_name", "")}
            for m in (all_tm.data or [])
            if m["id"] not in existing_ids
        ]

        owners_count = sum(1 for m in members if m["role"] == "owner")

        # current_role: role of the requesting user in this client, or None
        current_role = next(
            (m["role"] for m in members if m["member_id"] == user_id), None
        ) if user_id else None

        return {
            "members": members,
            "available": available,
            "is_owner": _is_owner(user_id, client_id),
            "current_role": current_role,          # "owner" | "member" | null
            "owners_count": owners_count,
            "can_claim": bool(user_id and owners_count == 0),
        }
    except Exception as e:
        print(f"get_client_members error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


# ── add_client_member ─────────────────────────────────────────────────────────
async def add_client_member(body: dict, user_id: Optional[str]):
    client_id = body.get("client_id")
    member_id = body.get("member_id")
    role = body.get("role", "member")

    if not client_id or not member_id:
        return JSONResponse({"error": "client_id et member_id requis"}, status_code=400)
    if role not in ("owner", "member"):
        return JSONResponse({"error": "role invalide"}, status_code=400)
    if not _is_owner(user_id, client_id):
        return JSONResponse({"error": "Seul un owner peut ajouter des membres"}, status_code=403)

    try:
        sb.table("client_members").upsert(
            {"client_id": client_id, "member_id": member_id, "role": role},
            on_conflict="client_id,member_id",
        ).execute()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    return {"added": True}


# ── remove_client_member ──────────────────────────────────────────────────────
async def remove_client_member(body: dict, user_id: Optional[str]):
    client_id = body.get("client_id")
    member_id = body.get("member_id")

    if not client_id or not member_id:
        return JSONResponse({"error": "client_id et member_id requis"}, status_code=400)

    self_removal = (member_id == user_id)
    if not self_removal and not _is_owner(user_id, client_id):
        return JSONResponse({"error": "Seul un owner peut retirer des membres"}, status_code=403)

    # Guard: don't remove last owner (applies to both self-removal and owner-removal)
    target = (
        sb.table("client_members").select("role")
        .eq("client_id", client_id).eq("member_id", member_id)
        .maybe_single().execute()
    )
    if target.data and target.data.get("role") == "owner" and _count_owners(client_id) <= 1:
        return JSONResponse({"error": "Impossible de quitter ce client : tu es le seul owner. Transfère l'ownership avant de partir."}, status_code=400)

    try:
        sb.table("client_members").delete().eq("client_id", client_id).eq("member_id", member_id).execute()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    return {"removed": True}


# ── set_member_role ───────────────────────────────────────────────────────────
async def set_member_role(body: dict, user_id: Optional[str]):
    client_id = body.get("client_id")
    member_id = body.get("member_id")
    role = body.get("role")

    if not client_id or not member_id or role not in ("owner", "member"):
        return JSONResponse({"error": "client_id, member_id et role (owner|member) requis"}, status_code=400)
    if not _is_owner(user_id, client_id):
        return JSONResponse({"error": "Seul un owner peut modifier les rôles"}, status_code=403)

    # Guard: don't demote last owner
    if role == "member":
        current = (
            sb.table("client_members").select("role")
            .eq("client_id", client_id).eq("member_id", member_id)
            .maybe_single().execute()
        )
        if current.data and current.data.get("role") == "owner" and _count_owners(client_id) <= 1:
            return JSONResponse({"error": "Impossible de rétrograder le dernier owner"}, status_code=400)

    try:
        sb.table("client_members").update({"role": role}).eq("client_id", client_id).eq("member_id", member_id).execute()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    return {"updated": True}


# ── claim_ownership ───────────────────────────────────────────────────────────
async def claim_ownership(body: dict, user_id: Optional[str]):
    """Lets any authenticated user become owner when the client has zero owners."""
    client_id = body.get("client_id")
    if not client_id:
        return JSONResponse({"error": "client_id requis"}, status_code=400)
    if not user_id:
        return JSONResponse({"error": "JWT requis pour claim_ownership"}, status_code=401)
    if _count_owners(client_id) > 0:
        return JSONResponse({"error": "Ce client a déjà un owner — demande-lui de te promouvoir"}, status_code=403)
    try:
        sb.table("client_members").upsert(
            {"client_id": client_id, "member_id": user_id, "role": "owner"},
            on_conflict="client_id,member_id",
        ).execute()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    _append_member_to_client_json(client_id, user_id)
    return {"claimed": True}


# ── upsert_task ───────────────────────────────────────────────────────────────
async def upsert_task(body: dict, user_id: Optional[str]):
    client_id = body.get("client_id")
    task      = body.get("task", {})
    if not client_id:
        return JSONResponse({"error": "client_id requis"}, status_code=400)
    await _assert_role(user_id, client_id, ["owner", "member"])
    task_id = task.get("id")
    if task_id and task_id > 0:
        sb.table("tasks").update({
            "title": task.get("title"), "prio": task.get("prio"),
            "status": task.get("status"), "assignee": task.get("assignee"),
            "blocker": task.get("blocker"), "note": task.get("note"),
            "due_date": task.get("due_date") or None,
            "scope": task.get("scope") or "internal",
            "last_modified_by": user_id,
            "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        }).eq("id", task_id).eq("client_id", client_id).execute()
    else:
        result = sb.table("tasks").insert({
            "client_id": client_id, "title": task.get("title"),
            "prio": task.get("prio") or "P2", "status": task.get("status") or "todo",
            "assignee": task.get("assignee") or "", "blocker": task.get("blocker") or None,
            "note": task.get("note") or None, "due_date": task.get("due_date") or None,
            "scope": task.get("scope") or "internal",
            "last_modified_by": user_id,
        }).execute()
        task = {**task, "id": result.data[0]["id"]}
    return JSONResponse({"task": task})


@observe(name="propose_cr_tasks", as_type="generation")
async def propose_cr_tasks(body: dict, user_id: Optional[str]):
    client_id = body.get("client_id")
    cr_text = (body.get("cr_text") or "").strip()
    if not client_id or not cr_text:
        return JSONResponse({"error": "client_id et cr_text requis"}, status_code=400)
    await _assert_role(user_id, client_id, ["owner", "member"])

    # Open tasks (not done)
    tasks_res = (
        sb.table("tasks")
        .select("id, title, status, assignee, prio, due_date")
        .eq("client_id", client_id)
        .neq("status", "done")
        .execute()
    )
    tasks_list = tasks_res.data or []

    # Client row: members JSON + brief context
    client_res = (
        sb.table("clients").select("context")
        .eq("id", client_id).maybe_single().execute()
    )
    client_data = client_res.data or {}

    # Smart Bees members — lus depuis les tables normalisées
    sb_members = []
    try:
        rows = (
            sb.table("client_members")
            .select("member_id, team_members(full_name, email)")
            .eq("client_id", client_id)
            .execute()
        )
        seen_initials = []
        for row in (rows.data or []):
            tm = row.get("team_members") or {}
            full_name = tm.get("full_name") or tm.get("email") or ""
            if not full_name:
                continue
            parts = full_name.strip().split()
            base = (parts[0][0] + parts[1][0]).upper() if len(parts) >= 2 else full_name[:2].upper()
            ini = base
            n = 2
            while ini in seen_initials:
                ini = base + str(n)
                n += 1
            seen_initials.append(ini)
            sb_members.append({"initials": ini, "name": full_name})
    except Exception:
        sb_members = []

    # Client-side contacts from the brief (equipe), if brief exists
    client_contacts = []
    try:
        ctx = client_data.get("context") or ""
        if ctx:
            brief = json.loads(ctx)
            _scope_map = {"interne": "internal", "externe": "external", "inconnu": "uncertain"}
            for m in brief.get("equipe", []):
                name = f"{m.get('prenom', '')} {m.get('nom') or ''}".strip()
                if name:
                    scope_raw = m.get("scope") or ""
                    client_contacts.append({
                        "name": name,
                        "role": m.get("role") or "",
                        "scope": _scope_map.get(scope_raw, scope_raw or "uncertain"),
                    })
    except Exception:
        pass

    sb_block = "\n".join(
        f"- {m.get('initials', '?')} : {m.get('name', '?')}"
        for m in sb_members
    ) or "(aucun membre SB enregistré pour ce client)"

    tasks_block = "\n".join(
        f"- ID {t['id']} : {t['title']}"
        + (f" [{t['status']}]" if t.get("status") else "")
        + (f" — {t['assignee']}" if t.get("assignee") else "")
        for t in tasks_list
    ) or "(aucune tâche ouverte)"

    contacts_block = ""
    if client_contacts:
        lines = "\n".join(
            f"- {c['name']}" + (f" ({c['role']})" if c.get("role") else "")
            for c in client_contacts
        )
        contacts_block = f"\nCONTACTS CÔTÉ CLIENT (actions leur étant assignées → scope=external) :\n{lines}\n"

    prompt = (
        "Tu es un assistant qui extrait les ACTIONS À FAIRE d'un compte-rendu de réunion "
        "pour une agence data/marketing.\n\n"
        f"ÉQUIPE SMART BEES (internes — seuls ceux-ci peuvent recevoir des tâches via 'assignee') :\n{sb_block}\n"
        + contacts_block
        + f"\nTÂCHES EN COURS (à mettre à jour si le CR les mentionne) :\n{tasks_block}\n\n"
        f"COMPTE-RENDU :\n{cr_text}\n\n"
        "Un CR mélange des actions à faire, des décisions, des options discutées et du contexte. "
        "Tu n'extrais QUE les ENGAGEMENTS : une action que quelqu'un s'est engagé à faire, "
        "ou une décision actée à exécuter.\n"
        "Priorité aux sections d'actions explicites ('Prochaines étapes', 'Actions', 'Next steps', 'TODO') : "
        "c'est la source de vérité. Dans le reste du CR, n'extrais une action que si c'est un engagement "
        "clair (« X va faire Y », « on doit Z »).\n\n"
        "NE SONT PAS des tâches : une option qu'on pèse (A/B/C), une recommandation non décidée, "
        "une contrainte, un constat ou une observation marché. Au mieux ça devient une 'note' sur une "
        "tâche liée — jamais une tâche en soi.\n\n"
        "Pour chaque engagement :\n"
        "- summary : résumé court et précis de l'action\n"
        "- match_type : 'update_existing' si elle correspond à une tâche en cours (fournis task_id), "
        "'new' pour une nouvelle tâche, 'uncertain' si tu n'es pas sûr\n"
        "- task_id : ID de la tâche existante si match_type=update_existing, null sinon — "
        "NE JAMAIS inventer un ID\n"
        "- scope : 'internal' si c'est du travail Smart Bees, 'external' si c'est une action "
        "côté client à suivre, 'uncertain' si non déterminable\n"
        "- fields : uniquement ce qui est dit EXPLICITEMENT dans le CR (null pour les autres) :\n"
        "  • title : titre de la tâche\n"
        "  • assignee : initiales SB UNIQUEMENT si la personne est dans la liste SB ci-dessus, null sinon\n"
        "  • prio : 'P1'/'P2'/'P3' si mentionné\n"
        "  • status : 'todo'/'inprogress'/'blocked'/'waiting'/'done' si le statut change\n"
        "  • due_date : date en YYYY-MM-DD si mentionnée\n"
        "  • note : contexte de reprise en 2-3 lignes max — pourquoi cette tâche existe, quel est le problème ou la décision d'origine, où on en est au moment du CR. Pas la phrase littérale du CR, une synthèse utile pour quelqu'un qui reprend après 3 semaines. null seulement si le titre se suffit vraiment à lui-même (tâches triviales type \"envoyer le doc X\")\n"
        "- confidence : 0.0 à 1.0\n"
        "- needs_clarification : true si une info cruciale manque\n"
        "- clarification_question : la question à poser, null si needs_clarification=false\n\n"
        "Règles strictes :\n"
        "1. Ne jamais inventer un task_id — si incertain → match_type='uncertain', task_id=null\n"
        "2. assignee = initiales SB uniquement depuis la liste fournie — null si inconnu\n"
        "3. Ne remplir fields que sur la base de ce qui est dit explicitement dans le CR\n"
        "4. Préférer needs_clarification=true plutôt que trancher avec peu d'info\n"
        "5. Au moindre doute « action ou simple discussion ? » → ne PAS créer de tâche\n"
        "6. Le nombre de tâches doit être stable que la section 'Prochaines étapes' soit présente "
        "ou non : les options et recommandations de la discussion ne sont pas des actions"
    )

    try:
        gemini = genai.GenerativeModel(
            model_name=GEMINI_FLASH,
            generation_config={
                "max_output_tokens": 4096,
                "response_mime_type": "application/json",
                "response_schema": _CR_PROPOSAL_SCHEMA,
            },
            safety_settings=_SAFETY_OFF,
        )
        response = gemini.generate_content(prompt)
        raw_text = _gemini_text(response)
    except Exception as e:
        return JSONResponse({"error": f"Erreur IA (CR) : {e}"}, status_code=502)

    # ── Langfuse (non bloquant) ────────────────────────────────────────────
    try:
        _um = getattr(response, "usage_metadata", None)
        _langfuse.update_current_generation(
            input=prompt[:3000],
            output=raw_text[:2000],
            model=GEMINI_FLASH,
            usage_details={"input": getattr(_um, "prompt_token_count", 0) if _um else 0,
                   "output": getattr(_um, "candidates_token_count", 0) if _um else 0},
            metadata={"client_id": body.get("client_id")},
        )
    except Exception:
        pass

    try:
        proposal = json.loads(raw_text)
    except Exception:
        return JSONResponse(
            {"error": "L'IA n'a pas retourné un JSON valide. Réessaie."},
            status_code=422,
        )

    # Invalidate task_ids that don't exist in this client's open tasks
    valid_ids = {t["id"] for t in tasks_list}
    for item in proposal.get("items", []):
        tid = item.get("task_id")
        if tid and tid not in valid_ids:
            item["match_type"] = "uncertain"
            item["task_id"] = None

    return JSONResponse(proposal)


# ── delete_task ────────────────────────────────────────────────────────────────
async def delete_task(body: dict, user_id: Optional[str]):
    task_id   = body.get("task_id")
    client_id = body.get("client_id")
    if not task_id or not client_id:
        return JSONResponse({"error": "task_id et client_id requis"}, status_code=400)
    await _assert_role(user_id, client_id, ["owner", "member"])
    sb.table("tasks").update({"last_modified_by": user_id}).eq("id", task_id).eq("client_id", client_id).execute()
    sb.table("tasks").delete().eq("id", task_id).eq("client_id", client_id).execute()
    return JSONResponse({"ok": True})


# ── weekly_digest ──────────────────────────────────────────────────────────────
@observe(name="weekly_digest", as_type="generation")
async def weekly_digest(body: dict, user_id: Optional[str]):
    if not user_id:
        return JSONResponse({"error": "JWT requis"}, status_code=401)

    force_refresh = body.get("force_refresh", False)
    CACHE_TTL_HOURS = 12

    # ── Check cache ──
    if not force_refresh:
        try:
            cached = (
                sb.table("digest_cache")
                .select("digest_text, generated_at, facts_hash")
                .eq("user_id", user_id)
                .single()
                .execute()
            )
            if cached.data:
                gen_at = datetime.fromisoformat(
                    cached.data["generated_at"].replace("Z", "+00:00")
                )
                age_hours = (datetime.now(timezone.utc) - gen_at).total_seconds() / 3600
                if age_hours < CACHE_TTL_HOURS:
                    return JSONResponse({
                        "digest": cached.data["digest_text"],
                        "empty": False,
                        "cached": True,
                        "generated_at": cached.data["generated_at"],
                    })
        except Exception:
            pass  # pas de cache, on génère

    # 1) Clients accessibles (même logique que /me)
    try:
        rows = (
            sb.table("client_members")
            .select("client_id, clients(id, name)")
            .eq("member_id", user_id)
            .execute()
        )
        clients = [
            {"id": r["client_id"], "name": (r.get("clients") or {}).get("name", "?")}
            for r in (rows.data or [])
        ]
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    if not clients:
        return JSONResponse({"digest": "Aucun client accessible cette semaine.", "empty": True})

    since = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    since_date = since[:10]
    facts_blocks = []

    for c in clients:
        cid = c["id"]
        # 2a) changements de la semaine
        try:
            hist = (
                sb.table("task_history")
                .select("task_id, action, field, old_value, new_value, changed_at")
                .eq("client_id", cid)
                .gte("changed_at", since)
                .order("changed_at", desc=False)
                .execute()
            ).data or []
        except Exception:
            hist = []

        # 2b) tâches ouvertes (non done) — retard et blocage
        try:
            open_tasks = (
                sb.table("tasks")
                .select("id, title, status, assignee, due_date, scope, blocker")
                .eq("client_id", cid)
                .neq("status", "done")
                .execute()
            ).data or []
        except Exception:
            open_tasks = []

        # 2c) tâches créées cette semaine (source : created_at, couvre aussi avant le trigger)
        try:
            new_tasks = (
                sb.table("tasks")
                .select("id, title, status")
                .eq("client_id", cid)
                .gte("created_at", since)
                .execute()
            ).data or []
        except Exception:
            new_tasks = []

        if not hist and not open_tasks and not new_tasks:
            continue

        titles = {t["id"]: t["title"] for t in open_tasks}
        missing = list({h["task_id"] for h in hist} - set(titles.keys()))
        if missing:
            try:
                extra = sb.table("tasks").select("id, title").in_("id", missing).execute().data or []
                for t in extra:
                    titles[t["id"]] = t["title"]
            except Exception:
                pass
        today = datetime.now(timezone.utc).date().isoformat()

        lines = [f"### Client : {c['name']}"]
        if new_tasks:
            lines.append("Nouvelles tâches cette semaine : " + ", ".join(t["title"] for t in new_tasks))
        if hist:
            lines.append("Changements cette semaine :")
            for h in hist:
                t_label = titles.get(h["task_id"], f"tâche #{h['task_id']}")
                if h["action"] == "deleted":
                    lines.append(f"- supprimée : tâche #{h['task_id']}")
                elif h["action"] != "created":
                    f = h.get("field")
                    if f != "note":
                        lines.append(f"- {t_label} : {f} {h.get('old_value')} → {h.get('new_value')}")

        # Notes récentes : toutes les tâches actives cette semaine (pas seulement celles avec field=note)
        active_task_ids = list({h["task_id"] for h in hist})
        if active_task_ids:
            try:
                note_rows = (
                    sb.table("tasks")
                    .select("id, title, note")
                    .in_("id", active_task_ids)
                    .execute()
                ).data or []
                for nr in note_rows:
                    note_txt = nr.get("note") or ""
                    if not note_txt.strip():
                        continue
                    recent = _recent_note_entries(note_txt, since_date)
                    if recent:
                        for entry in recent:
                            lines.append(f"- note récente sur « {nr['title']} » : {entry}")
                    # Toujours inclure les 2 dernières entrées de note pour le contexte,
                    # même si elles sont hors fenêtre de 7 jours
                    all_entries = _recent_note_entries(note_txt, "2000-01-01")
                    if all_entries:
                        latest = all_entries[-2:]
                        for entry in latest:
                            if entry not in recent:
                                lines.append(f"- contexte sur « {nr['title']} » : {entry}")
            except Exception as _note_err:
                print(f"digest note fetch error (non bloquant): {_note_err}")

        # Sessions récentes : résumés de conversations de la semaine
        try:
            session_chunks = (
                sb.table("document_chunks")
                .select("source_name, chunk_text, last_indexed_at")
                .eq("client_id", cid)
                .eq("source_type", "session")
                .gte("last_indexed_at", since)
                .order("last_indexed_at", desc=True)
                .limit(5)
                .execute()
            ).data or []
            if session_chunks:
                lines.append("Points abordés dans les sessions de la semaine :")
                for sc in session_chunks:
                    summary = sc["chunk_text"][:300].replace('\n', ' ').strip()
                    lines.append(f"- {sc['source_name']} : {summary}")
        except Exception:
            pass

        newly_blocked_ids = {
            h["task_id"] for h in hist
            if h.get("field") == "status" and h.get("new_value") in ("blocked", "waiting")
        }

        ours = [t for t in open_tasks if t.get("scope") != "external"]
        external = [t for t in open_tasks if t.get("scope") == "external"]

        late_new = [t for t in ours if t.get("due_date") and since_date <= t["due_date"] < today]
        still_waiting = [
            t for t in ours
            if (t.get("status") in ("blocked", "waiting") and t["id"] not in newly_blocked_ids)
            or (t.get("due_date") and t["due_date"] < since_date)
        ]
        ext_watch = [t for t in external
                     if (t.get("due_date") and t["due_date"] < today)
                     or t.get("status") in ("blocked", "waiting")]

        # Raisons de blocage — nouvellement bloquées cette semaine
        task_blocker = {t["id"]: (t.get("blocker") or "").strip() for t in open_tasks}
        for tid in newly_blocked_ids:
            b = task_blocker.get(tid, "")
            if b:
                lines.append(f"- bloqué : {titles.get(tid, f'tâche #{tid}')} — {b}")
        # Raisons de blocage — toujours en attente (status blocked/waiting)
        for t in still_waiting:
            if t.get("status") in ("blocked", "waiting"):
                b = (t.get("blocker") or "").strip()
                if b:
                    lines.append(f"- bloqué : {t['title']} — {b}")

        if late_new:
            lines.append("Échéance dépassée cette semaine : " + ", ".join(t["title"] for t in late_new))
        if still_waiting:
            lines.append("Toujours en attente : " + ", ".join(t["title"] for t in still_waiting))
        if ext_watch:
            lines.append("À suivre côté client : " + ", ".join(t["title"] for t in ext_watch))
        facts_blocks.append("\n".join(lines))

    if not facts_blocks:
        return JSONResponse({"digest": "Rien n'a bougé cette semaine sur tes clients.", "empty": True})

    facts = "\n\n".join(facts_blocks)

    # 3) Mise en mots par Gemini Flash
    try:
        gemini = genai.GenerativeModel(
            model_name=GEMINI_FLASH,
            system_instruction=(
                "Tu es un assistant qui rédige un récap hebdomadaire d'activité pour une agence "
                "data/marketing. Ton orienté business : ce qui a avancé, ce qui s'est débloqué, "
                "les nouveaux points bloquants, ce qui traîne. Factuel, concis, pas de remplissage. "
                "Tu écris en 'on' (première personne du pluriel)."
            ),
            generation_config={"max_output_tokens": 8000},
            safety_settings=_SAFETY_OFF,
        )
        prompt = (
            "Voici les faits bruts de la semaine, par client. Rédige un récap clair et court, "
            "groupé par client, qui met en avant les avancées, les déblocages et les points bloquants. "
            "Ne liste pas mécaniquement chaque changement de champ : synthétise en langage naturel. "
            "N'écris JAMAIS les valeurs de statut brutes (todo, inprogress, blocked, waiting, done) — "
            "utilise des formulations françaises : « démarré », « terminé », « bloqué », "
            "« en attente de [raison] », « en cours ». Les noms de tâches entre guillemets français (« »), pas droits. "
            "Mentionne TOUJOURS les tâches terminées et les nouvelles tâches créées : ce sont les "
            "signaux les plus importants, ne les omets jamais. "
            "Maximum 2-3 lignes par tâche. La première ligne résume le changement de statut. "
            "Les lignes suivantes intègrent le CONTENU des notes récentes (décisions, retours, contexte). "
            "Les notes apportent le 'pourquoi' et le 'quoi' — ne les omets jamais au profit du simple statut. "
            "Si une note contient une décision ou un retour d'une personne, cite-la. "
            "Tu peux regrouper le reste, mais ne supprime "
            "pas une avancée ou un blocage. Si un client n'a vraiment rien (aucune ligne de faits), ne le mentionne pas.\n\n"
            "FORMAT IMPÉRATIF de ta réponse :\n"
            "- Le nom de chaque client sur une ligne seule, sans tiret ni ponctuation autour.\n"
            "- En dessous, chaque point sur sa propre ligne, préfixé par '– ' (tiret long + espace).\n"
            "- Une ligne vide entre deux clients.\n"
            "- Pas de titre général, pas d'introduction, pas de conclusion.\n"
            "- Si un client n'a rien de notable, ne le mentionne pas du tout.\n"
            "- Les tâches terminées et créées cette semaine doivent apparaître explicitement.\n"
            "- « Échéance dépassée cette semaine » = nouveaux retards, à signaler en priorité.\n"
            "- « Toujours en attente » = tâches qui stagnent depuis avant cette semaine : mentionne-les "
            "brièvement, séparément des nouveautés, sans en faire un drame.\n"
            "- « À suivre côté client » = actions externes qu'on surveille sans en être responsables : "
            "présente-les comme du suivi, pas comme nos retards.\n"
            "- Les « note récente sur … » contiennent des décisions, retours, ou contexte clé : "
            "intègre-les naturellement dans le récap de la tâche concernée (pas comme une ligne séparée).\n"
            "- Les « Points abordés dans les sessions » sont des résumés de conversations : "
            "mentionne les décisions ou avancées importantes, pas les détails opérationnels.\n\n"
            + facts
        )
        response = gemini.generate_content(prompt)
        digest_text = _gemini_text(response)
        finish = response.candidates[0].finish_reason
        finish_name = finish.name if hasattr(finish, "name") else str(finish)
        if finish_name == "MAX_TOKENS":
            digest_text += "\n\n⚠️ Récap tronqué (trop d'activité cette semaine). Réessaie ou consulte les tâches directement."
    except Exception as e:
        return JSONResponse({"error": f"Erreur IA (digest) : {e}"}, status_code=502)

    # 4) log coût (non bloquant)
    try:
        usage = getattr(response, "usage_metadata", None)
        _in  = getattr(usage, "prompt_token_count", 0) if usage else 0
        _out = getattr(usage, "candidates_token_count", 0) if usage else 0
        sb.table("usage_logs").insert({
            "user_id": user_id,
            "action": "weekly_digest",
            "model": GEMINI_FLASH,
            "tokens_input": _in,
            "tokens_output": _out,
            "cost_usd": calculate_cost(GEMINI_FLASH, {"input_tokens": _in, "output_tokens": _out}),
        }).execute()
    except Exception as e:
        print(f"usage_logs digest (non bloquant): {e}")

    # ── Langfuse (non bloquant) ────────────────────────────────────────────
    try:
        _langfuse.update_current_generation(
            input=prompt[:3000],
            output=digest_text[:2000],
            model=GEMINI_FLASH,
            usage_details={"input": _in, "output": _out},
            metadata={"user_id": user_id, "scope": "all_clients"},
        )
    except Exception:
        pass

    # ── Sauvegarder en cache ──
    try:
        fhash = hashlib.md5(facts.encode()).hexdigest()[:16]
        sb.table("digest_cache").upsert({
            "user_id": user_id,
            "digest_text": digest_text,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "facts_hash": fhash,
        }).execute()
    except Exception as _cache_err:
        print(f"digest cache write error (non bloquant): {_cache_err}")

    return JSONResponse({"digest": digest_text, "empty": False, "cached": False})


# ── delete_client ──────────────────────────────────────────────────────────────
async def create_client(body: dict, user_id: Optional[str]):
    """Creates a client row + initial client_members rows atomically. Requires JWT."""
    if not user_id:
        return JSONResponse({"error": "JWT requis pour créer un client"}, status_code=401)
    name = (body.get("name") or "").strip()
    members_json = body.get("members_json", [])
    member_rows  = body.get("member_rows", [])   # [{member_id, role}]
    if not name:
        return JSONResponse({"error": "Le nom du client est requis"}, status_code=400)
    try:
        res = sb.table("clients").insert({
            "name": name,
            "members": json.dumps(members_json),
        }).execute()
        if not res.data:
            return JSONResponse({"error": "Erreur création client"}, status_code=500)
        client = res.data[0]
        rows = [{"client_id": client["id"], "member_id": r["member_id"], "role": r["role"]}
                for r in member_rows if r.get("member_id")]
        if rows:
            sb.table("client_members").insert(rows).execute()
        return JSONResponse({"client": client})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


async def delete_client(body: dict, user_id: Optional[str]):
    client_id = body.get("client_id")
    if not client_id:
        return JSONResponse({"error": "client_id requis"}, status_code=400)
    await _assert_role(user_id, client_id, ["owner"])
    sb.table("clients").delete().eq("id", client_id).execute()
    return JSONResponse({"ok": True})


# ── create_invitation ─────────────────────────────────────────────────────────
async def create_invitation(body: dict, user_id: Optional[str]):
    client_id     = body.get("client_id")
    invited_email = body.get("invited_email", "").strip().lower()
    role          = body.get("role", "member")

    if not client_id or not invited_email:
        return JSONResponse({"error": "client_id et invited_email requis"}, status_code=400)
    if not user_id:
        return JSONResponse({"error": "JWT requis pour créer une invitation"}, status_code=401)
    if role not in ("owner", "member"):
        return JSONResponse({"error": "Rôle invalide — valeurs acceptées : owner, member"}, status_code=400)

    await _assert_role(user_id, client_id, ["owner"])

    inv = sb.table("client_invitations").insert({
        "client_id":     client_id,
        "created_by":    user_id,
        "invited_email": invited_email,
        "role":          role,
    }).execute()

    token      = inv.data[0]["token"]
    base_url   = os.getenv("FRONTEND_URL", "https://khadija-benayed.github.io/clientchat_v2").rstrip('/')
    invite_url = f"{base_url}/#/join/{token}"
    return JSONResponse({"token": token, "url": invite_url, "expires_at": inv.data[0]["expires_at"]})


# ── join_client_via_token ─────────────────────────────────────────────────────
async def join_client_via_token(body: dict, user_id: Optional[str]):
    token = body.get("token", "").strip()
    if not token:
        return JSONResponse({"error": "Token requis"}, status_code=400)
    if not user_id:
        return JSONResponse({"error": "JWT requis"}, status_code=401)

    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    inv = sb.table("client_invitations") \
        .select("*") \
        .eq("token", token) \
        .is_("used_at", "null") \
        .gt("expires_at", now_iso) \
        .maybe_single().execute()

    if not inv.data:
        return JSONResponse({"error": "Invitation invalide ou expirée"}, status_code=403)

    inv_data = inv.data

    # Vérifier que l'email correspond
    try:
        user_obj   = sb.auth.admin.get_user(user_id)
        user_email = user_obj.user.email.lower()
    except Exception:
        return JSONResponse({"error": "Impossible de vérifier l'identité"}, status_code=503)

    if user_email != inv_data["invited_email"]:
        return JSONResponse(
            {"error": "Ce lien d'invitation ne correspond pas à votre compte"},
            status_code=403,
        )

    # Marquer l'invitation comme utilisée de façon atomique (fix TOCTOU)
    # Le filtre .is_("used_at", "null") garantit qu'un seul appel concurrent réussit.
    claimed = sb.table("client_invitations").update({
        "used_at": now_iso,
        "used_by": user_id,
    }).eq("id", inv_data["id"]).is_("used_at", "null").execute()

    if not claimed.data:
        return JSONResponse({"error": "Invitation déjà utilisée"}, status_code=409)

    # Ajouter le membre (upsert — idempotent si doublon résiduel)
    try:
        sb.table("client_members").upsert({
            "client_id": inv_data["client_id"],
            "member_id": user_id,
            "role":      inv_data["role"],
        }, on_conflict="client_id,member_id").execute()
    except Exception:
        # Rollback: libérer le token pour permettre une nouvelle tentative
        sb.table("client_invitations").update({"used_at": None, "used_by": None}).eq("id", inv_data["id"]).execute()
        return JSONResponse({"error": "Erreur lors de l'ajout du membre, réessaie"}, status_code=500)

    _append_member_to_client_json(inv_data["client_id"], user_id)

    client = sb.table("clients").select("*").eq("id", inv_data["client_id"]).maybe_single().execute()
    if not client.data:
        return JSONResponse({"error": "L'espace client n'existe plus"}, status_code=410)
    return JSONResponse({"client": client.data})


# ── sync_drive ────────────────────────────────────────────────────────────────
async def sync_drive(body: dict, request: Request):
    folder_id = body.get("folder_id")
    client_id = body.get("client_id")
    resume = body.get("resume", False)
    incremental = body.get("incremental", True)  # default True — frontend may omit it (cache)

    if not folder_id or not client_id:
        return JSONResponse({"error": "folder_id et client_id requis"}, status_code=400)
    await _assert_role(getattr(request.state, "user_id", None), client_id, ["owner", "member"])

    try:
        drive, sa_email = get_drive_service()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    total_ref = [0]
    all_files = _list_files_recursive(drive, folder_id, set(), 500, total_ref)

    if not all_files:
        return JSONResponse({"error": f"Aucun fichier trouvé. Vérifie le partage avec {sa_email}."}, status_code=404)

    all_files.sort(key=lambda f: (_type_priority(f["mimeType"]), -_parse_modified(f).timestamp()))
    files_to_process = all_files[:200]
    total = len(files_to_process)

    # Preload indexed state — paginated to bypass PostgREST 1000-row default
    existing_ids: set = set()           # source_ids currently in DB
    indexed_at: dict = {}               # incremental: {source_id: last_indexed datetime}
    if client_id:
        try:
            PAGE = 1000
            offset = 0
            while True:
                res = (sb.table("document_chunks")
                       .select("source_id, last_indexed_at")
                       .eq("client_id", client_id)
                       .range(offset, offset + PAGE - 1)
                       .execute())
                rows = res.data or []
                for r in rows:
                    sid, lat = r.get("source_id"), r.get("last_indexed_at")
                    if not sid:
                        continue
                    existing_ids.add(sid)
                    if lat:
                        try:
                            dt = datetime.fromisoformat(lat.replace("Z", "+00:00"))
                            if sid not in indexed_at or dt > indexed_at[sid]:
                                indexed_at[sid] = dt
                        except Exception:
                            pass
                if len(rows) < PAGE:
                    break
                offset += PAGE
        except Exception as e:
            print(f"sync_drive state preload error (non bloquant): {e}")

    # Preload permanently-ignored file IDs so _is_cached() skips them without
    # retrying a 150 s timeout per file on every sync. Only PERMANENT reasons
    # are loaded — transient ones ('timeout', 'error') stay retriable.
    permanently_ignored_ids: set = set()
    if client_id:
        try:
            PAGE = 1000
            offset = 0
            while True:
                res = (sb.table("sync_ignored")
                       .select("source_id")
                       .eq("client_id", client_id)
                       .in_("reason", list(_PERM_IGNORE_REASONS))
                       .range(offset, offset + PAGE - 1)
                       .execute())
                rows = res.data or []
                for r in rows:
                    if r.get("source_id"):
                        permanently_ignored_ids.add(r["source_id"])
                if len(rows) < PAGE:
                    break
                offset += PAGE
        except Exception as e:
            print(f"sync_drive perm_ignored preload error (non bloquant): {e}")

    mode = "resume" if resume else ("incremental" if incremental else "full")
    print(f"sync_drive: mode={mode} total={len(files_to_process)} existing={len(existing_ids)} indexed_at={len(indexed_at)}")

    # Purge chunks for files deleted from Drive (source_id in DB but not in Drive).
    # Only safe when the Drive listing is complete — if it was capped at 500 files,
    # files beyond the cap would appear as ghosts and be incorrectly deleted.
    drive_ids = {f["id"] for f in all_files}
    listing_complete = len(all_files) < 500
    ghost_ids = (existing_ids - drive_ids) if listing_complete else set()
    purged_count = 0
    if ghost_ids and client_id:
        try:
            sb.table("document_chunks").delete().eq("client_id", client_id).in_("source_id", list(ghost_ids)).execute()
            purged_count = len(ghost_ids)
            print(f"sync_drive: purged {purged_count} deleted Drive file(s) from DB")
        except Exception as e:
            print(f"sync_drive: ghost cleanup error (non bloquant): {e}")

    state_key = f"{client_id}|{folder_id}"
    _sync_state[state_key] = {
        "total": total, "processed": 0, "ok": 0,
        "cached": 0, "skipped": 0, "errors": 0, "purged": purged_count, "done": False,
    }

    async def generate():
        processed = 0
        ok = 0
        cached = 0
        skipped = 0
        errors = 0
        BATCH_SIZE = 5
        loop = asyncio.get_running_loop()

        def _upd(**kw):
            _sync_state[state_key].update({"processed": processed, "ok": ok,
                                           "cached": cached, "errors": errors, **kw})

        def _ignore(file_id, file_name, reason):
            try:
                sb.table('sync_ignored').upsert({
                    'source_id':   file_id,
                    'client_id':   client_id,
                    'source_name': file_name,
                    'reason':      reason,
                    'ignored_at':  time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
                }, on_conflict='source_id').execute()
            except Exception:
                pass

        for i in range(0, total, BATCH_SIZE):
            batch = files_to_process[i:i + BATCH_SIZE]

            # Decide which files to skip (cached) vs download
            def _is_cached(f: dict) -> bool:
                fid = f["id"]
                # Fichier définitivement non-indexable : skip sans téléchargement
                if fid in permanently_ignored_ids:
                    return True
                if resume and fid in existing_ids:
                    return True
                if incremental and fid in indexed_at:
                    return _parse_modified(f) <= indexed_at[fid]
                return False

            for f in [f for f in batch if _is_cached(f)]:
                processed += 1
                cached += 1
                _upd()
                yield f"data: {json.dumps({'file': f['name'], 'status': 'cached', 'progress': processed, 'total': total})}\n\n"

            to_fetch = [f for f in batch if not _is_cached(f)]
            if not to_fetch:
                continue

            # Parallel download with heartbeats every 5 s + 90 s hard timeout per batch
            # Per-file timeout (150 s each) — each file has its own budget.
            # A slow file (large scanned PDF + Claude Vision) gets its full time
            # without blocking or penalising the others in the batch.
            MAX_FILE_WAIT = 150
            futs = [
                asyncio.ensure_future(loop.run_in_executor(
                    None, export_drive_file, f["id"], f["name"], f["mimeType"]
                ))
                for f in to_fetch
            ]
            deadlines = {fut: loop.time() + MAX_FILE_WAIT for fut in futs}
            timed_out_ids: set = set()
            pending = set(futs)
            while pending:
                done_set, pending = await asyncio.wait(pending, timeout=5.0)
                now = loop.time()
                for i, fut in enumerate(futs):
                    if fut in pending and now >= deadlines[fut]:
                        fut.cancel()
                        pending.discard(fut)
                        timed_out_ids.add(to_fetch[i]["id"])
                if pending:
                    yield f"data: {json.dumps({'status': 'heartbeat', 'progress': processed, 'total': total})}\n\n"

            results = []
            for fut in futs:
                try:
                    results.append(None if fut.cancelled() else fut.result())
                except Exception:
                    results.append(_EXPORT_TRANSIENT_ERR)

            for f, result in zip(to_fetch, results):
                processed += 1
                _upd()

                if f["id"] in timed_out_ids:
                    errors += 1
                    _upd()
                    _ignore(f['id'], f['name'], 'timeout')
                    yield f"data: {json.dumps({'file': f['name'], 'status': 'timeout', 'progress': processed, 'total': total})}\n\n"
                    continue

                if result is _EXPORT_TRANSIENT_ERR:
                    errors += 1
                    _upd()
                    _ignore(f['id'], f['name'], 'error')
                    yield f"data: {json.dumps({'file': f['name'], 'status': 'error', 'progress': processed, 'total': total})}\n\n"
                    continue

                if result is None:
                    skipped += 1
                    _ignore(f['id'], f['name'], 'skipped')
                    yield f"data: {json.dumps({'file': f['name'], 'mimeType': f['mimeType'], 'status': 'skipped', 'progress': processed, 'total': total})}\n\n"
                    continue

                try:
                    content = result["content"]
                    if not content.strip():
                        _ignore(f['id'], f['name'], 'empty')
                        yield f"data: {json.dumps({'file': f['name'], 'status': 'empty', 'progress': processed, 'total': total})}\n\n"
                        continue

                    is_csv = result["type"] in ("csv", "sheet")
                    chunks = chunk_csv(content) if is_csv else chunk_text(content)
                    if not chunks:
                        _ignore(f['id'], f['name'], 'empty')
                        yield f"data: {json.dumps({'file': f['name'], 'status': 'empty', 'progress': processed, 'total': total})}\n\n"
                        continue

                    # Prefix = "filename [dd/mm/yyyy]\n" — same as index_source.
                    _fdate = ""
                    if f.get("modifiedTime"):
                        try:
                            _fdt = datetime.fromisoformat(f["modifiedTime"].replace("Z", "+00:00"))
                            _fdate = f" [{_fdt.strftime('%d/%m/%Y')}]"
                        except Exception:
                            pass
                    prefixed_chunks = [f["name"][:60] + _fdate + "\n" + c for c in chunks]
                    # Heartbeat during embedding (same pattern as downloads)
                    embed_task = asyncio.ensure_future(
                        loop.run_in_executor(_EMBED_EXECUTOR, embed_texts, prefixed_chunks)
                    )
                    while not embed_task.done():
                        await asyncio.wait({embed_task}, timeout=5.0)
                        if not embed_task.done():
                            yield f"data: {json.dumps({'status': 'heartbeat', 'progress': processed, 'total': total})}\n\n"
                    embeddings = embed_task.result()

                    try:
                        del_q = sb.table("document_chunks").delete()
                        if client_id:
                            del_q = del_q.eq("client_id", client_id)
                        else:
                            del_q = del_q.is_("client_id", "null")
                        del_q.eq("source_id", f["id"]).execute()
                    except Exception:
                        pass

                    now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
                    rows = [{
                        "client_id": client_id or None,
                        "source_type": result["type"],
                        "source_name": f["name"],
                        "source_id": f["id"],
                        "chunk_text": chunk,
                        "embedding": emb,
                        "last_indexed_at": now,
                        **({"drive_modified_at": f["modifiedTime"]} if f.get("modifiedTime") else {}),
                    } for chunk, emb in zip(chunks, embeddings)]

                    _sb_insert("document_chunks", rows)
                    ok += 1
                    _upd()
                    yield f"data: {json.dumps({'file': f['name'], 'status': 'ok', 'chunks': len(chunks), 'progress': processed, 'total': total})}\n\n"

                except Exception as e:
                    err_str = str(e).lower()
                    reason = (
                        'ineligible_ai'  if 'ineligible' in err_str or 'generative ai' in err_str
                        else 'export_error' if '403' in err_str or 'export' in err_str
                        else 'timeout'    if 'timeout' in err_str or 'cancel' in err_str
                        else 'error'
                    )
                    _ignore(f['id'], f['name'], reason)
                    errors += 1
                    _upd()
                    yield f"data: {json.dumps({'file': f['name'], 'status': 'ignored', 'reason': reason, 'progress': processed, 'total': total})}\n\n"

        _sync_state[state_key]["done"] = True
        yield f"data: {json.dumps({'status': 'done', 'total': total, 'ok': ok, 'cached': cached, 'errors': errors, 'purged': purged_count})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


# ── summarize_with_llm ───────────────────────────────────────────────────────
def summarize_with_llm(text: str) -> str:
    """Summarize email thread text via Gemini Flash. Returns 'SKIP' if no business info found."""
    try:
        gemini = genai.GenerativeModel(
            model_name=GEMINI_FLASH,
            system_instruction=(
                "Tu es un assistant qui extrait les informations métier d'échanges email.\n"
                "Extrait UNIQUEMENT : décisions prises, chiffres clés, engagements pris, points bloquants, actions à faire.\n"
                "Format : liste à tirets, 5 points max, sois factuel.\n"
                "Ne mentionne jamais les noms des expéditeurs ni les adresses email.\n"
                "Si le thread ne contient aucune information métier pertinente, réponds uniquement : SKIP"
            ),
            generation_config={"max_output_tokens": 400},
            safety_settings=_SAFETY_OFF,
        )
        response = gemini.generate_content(text)
        return _gemini_text(response).strip()
    except ValueError:
        # _gemini_text raises ValueError when Gemini blocks or returns no content — treat as SKIP
        return "SKIP"
    except Exception as e:
        # API quota, network, auth errors — propagate so the caller counts this as an error
        print(f"summarize_with_llm error: {e}")
        raise


# ── sync_emails ───────────────────────────────────────────────────────────────
async def sync_emails(body: dict, request: Request):
    client_id = body.get("client_id")
    label_name = body.get("label_name", "").strip()
    days_back = max(1, min(90, int(body.get("days_back", 7))))
    user_id = getattr(request.state, "user_id", None)

    if not client_id or not label_name:
        return JSONResponse({"error": "client_id et label_name requis"}, status_code=400)
    if not user_id:
        return JSONResponse({"error": "JWT requis"}, status_code=401)

    # Récupérer les membres opt-in Gmail pour ce client
    try:
        rows = (
            sb.table("client_members")
            .select("member_id, team_members(id, email, gmail_sync_enabled)")
            .eq("client_id", client_id)
            .execute()
        )
        opted_in_emails = []
        for row in (rows.data or []):
            tm = row.get("team_members") or {}
            if tm.get("gmail_sync_enabled") and tm.get("email"):
                opted_in_emails.append(tm["email"])
    except Exception as e:
        return JSONResponse({"error": f"Erreur récupération membres : {e}"}, status_code=500)

    async def _no_members():
        yield f"data: {json.dumps({'status': 'done', 'total': 0, 'ok': 0, 'skipped': 0, 'errors': 0, 'message': 'Aucun membre avec Gmail activé pour ce client'})}\n\n"

    if not opted_in_emails:
        return StreamingResponse(_no_members(), media_type="text/event-stream", headers={
            "Cache-Control": "no-cache", "X-Accel-Buffering": "no",
        })

    # Calculer la date de coupure (format Gmail : YYYY/MM/DD)
    cutoff_ts = int(time.time()) - days_back * 86400
    cutoff_date = datetime.utcfromtimestamp(cutoff_ts).strftime('%Y/%m/%d')

    async def generate():
        loop = asyncio.get_running_loop()
        seen_thread_ids: set = set()
        all_threads: list = []
        ok = 0
        skipped = 0
        errors = 0

        # ── Collecte des threads depuis toutes les boîtes opt-in ──────────────
        for email in opted_in_emails:
            try:
                gmail = await loop.run_in_executor(None, get_gmail_service, email)
                query = f'label:"{label_name}" after:{cutoff_date}'
                page_token = None
                while True:
                    params: dict = {"userId": "me", "q": query, "maxResults": 100}
                    if page_token:
                        params["pageToken"] = page_token
                    result = await loop.run_in_executor(
                        None,
                        lambda p=params, g=gmail: g.users().threads().list(**p).execute()
                    )
                    for t in result.get("threads", []):
                        if t["id"] not in seen_thread_ids:
                            seen_thread_ids.add(t["id"])
                            all_threads.append({"id": t["id"], "gmail": gmail})
                    page_token = result.get("nextPageToken")
                    if not page_token:
                        break
            except Exception as e:
                print(f"sync_emails: Gmail list error for {email}: {e}")
                yield f"data: {json.dumps({'status': 'heartbeat', 'progress': 0, 'total': 0})}\n\n"

        total = len(all_threads)
        if total == 0:
            yield f"data: {json.dumps({'status': 'done', 'total': 0, 'ok': 0, 'skipped': 0, 'errors': 0})}\n\n"
            return

        processed = 0

        for item in all_threads:
            thread_id = item["id"]
            gmail = item["gmail"]
            processed += 1

            try:
                # Télécharger le thread complet avec heartbeat
                fetch_task = asyncio.ensure_future(
                    loop.run_in_executor(
                        None,
                        lambda tid=thread_id, g=gmail: g.users().threads().get(
                            userId="me", id=tid, format="full"
                        ).execute()
                    )
                )
                while not fetch_task.done():
                    await asyncio.wait({fetch_task}, timeout=5.0)
                    if not fetch_task.done():
                        yield f"data: {json.dumps({'status': 'heartbeat', 'progress': processed, 'total': total})}\n\n"
                thread_data = fetch_task.result()

                messages = thread_data.get("messages", [])

                # Extraire le sujet depuis le premier message
                subject = "Sans sujet"
                if messages:
                    for h in messages[0].get("payload", {}).get("headers", []):
                        if h.get("name", "").lower() == "subject":
                            subject = h["value"][:100]
                            break

                # Construire le texte du thread (corps de tous les messages)
                parts_text = []
                for msg in messages:
                    msg_date = ""
                    for h in msg.get("payload", {}).get("headers", []):
                        if h.get("name", "").lower() == "date":
                            msg_date = h["value"][:30]
                            break
                    body_text = _extract_email_text(msg.get("payload", {}))
                    if body_text:
                        parts_text.append(f"[{msg_date}]\n{body_text}" if msg_date else body_text)

                thread_text = "\n\n---\n\n".join(parts_text)[:8000]
                if not thread_text.strip():
                    skipped += 1
                    yield f"data: {json.dumps({'thread_id': thread_id, 'subject': subject, 'status': 'skipped', 'progress': processed, 'total': total})}\n\n"
                    continue

                # Résumer via Gemini Flash avec heartbeat
                summarize_task = asyncio.ensure_future(
                    loop.run_in_executor(None, summarize_with_llm, thread_text)
                )
                while not summarize_task.done():
                    await asyncio.wait({summarize_task}, timeout=5.0)
                    if not summarize_task.done():
                        yield f"data: {json.dumps({'status': 'heartbeat', 'progress': processed, 'total': total})}\n\n"
                summary = summarize_task.result()

                if summary.strip().upper() == "SKIP":
                    skipped += 1
                    yield f"data: {json.dumps({'thread_id': thread_id, 'subject': subject, 'status': 'skipped', 'progress': processed, 'total': total})}\n\n"
                    continue

                # Embed + stocker dans document_chunks (jamais le corps brut)
                embedding = (await loop.run_in_executor(_EMBED_EXECUTOR, embed_texts, [summary]))[0]
                now_str = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

                try:
                    sb.table("document_chunks").delete().eq("client_id", client_id).eq("source_id", thread_id).eq("source_type", "email_summary").execute()
                except Exception:
                    pass

                sb.table("document_chunks").insert({
                    "client_id": client_id,
                    "source_type": "email_summary",
                    "source_name": f"Email — {subject[:60]}",
                    "source_id": thread_id,
                    "chunk_text": summary,
                    "embedding": embedding,
                    "last_indexed_at": now_str,
                }).execute()

                ok += 1
                yield f"data: {json.dumps({'thread_id': thread_id, 'subject': subject, 'status': 'summarized', 'progress': processed, 'total': total})}\n\n"

            except Exception as e:
                errors += 1
                print(f"sync_emails: error thread {thread_id}: {e}")
                yield f"data: {json.dumps({'thread_id': thread_id, 'status': 'error', 'error': str(e)[:200], 'progress': processed, 'total': total})}\n\n"

        # Purger les threads hors fenêtre de dates — fait APRÈS indexation pour ne pas
        # perdre les données existantes si la sync est interrompue en cours de route.
        if seen_thread_ids:
            try:
                sb.table("document_chunks").delete()\
                    .eq("client_id", client_id)\
                    .eq("source_type", "email_summary")\
                    .not_.in_("source_id", list(seen_thread_ids))\
                    .execute()
            except Exception as e:
                print(f"sync_emails: purge stale summaries error (non bloquant): {e}")

        yield f"data: {json.dumps({'status': 'done', 'total': total, 'ok': ok, 'skipped': skipped, 'errors': errors})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


# ── update_gmail_sync ─────────────────────────────────────────────────────────
async def update_gmail_sync(body: dict, user_id: Optional[str]):
    enabled = body.get("enabled")
    if enabled is None:
        return JSONResponse({"error": "enabled requis (bool)"}, status_code=400)
    if not user_id:
        return JSONResponse({"error": "JWT requis"}, status_code=401)
    try:
        sb.table("team_members").update({"gmail_sync_enabled": bool(enabled)}).eq("id", user_id).execute()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    return {"updated": True}


# ── web search stream helper ─────────────────────────────────────────────────
async def _chat_web_search_stream(system: str, message: str, user_id: Optional[str], client_id: str, lf_trace=None):
    loop = asyncio.get_running_loop()
    q: asyncio.Queue = asyncio.Queue()
    cancel_event = threading.Event()

    def _sync():
        try:
            gm = genai.GenerativeModel(
                model_name=GEMINI_FLASH,
                system_instruction=system,
                generation_config={"max_output_tokens": 3000},
                tools=[genai.protos.Tool(
                    google_search_retrieval=genai.protos.GoogleSearchRetrieval()
                )],
                safety_settings=_SAFETY_OFF,
            )
            contents = [{"role": "user", "parts": [message]}]
            resp = gm.generate_content(contents, stream=True, request_options={"timeout": 120})
            for chunk in resp:
                if cancel_event.is_set():
                    break
                try:
                    t = chunk.text
                    if t:
                        loop.call_soon_threadsafe(q.put_nowait, ("tok", t))
                except Exception:
                    pass
            if not cancel_event.is_set():
                loop.call_soon_threadsafe(q.put_nowait, ("done", resp))
        except Exception as exc:
            if not cancel_event.is_set():
                loop.call_soon_threadsafe(q.put_nowait, ("err", str(exc)))

    threading.Thread(target=_sync, daemon=True).start()

    try:
        accumulated = ""
        stream_resp = None

        while True:
            try:
                kind, data = await asyncio.wait_for(q.get(), timeout=120.0)
            except asyncio.TimeoutError:
                yield f"data: {json.dumps({'type': 'error', 'message': 'Timeout IA'})}\n\n"
                return
            if kind == "tok":
                accumulated += data
                yield f"data: {json.dumps({'type': 'token', 'text': data})}\n\n"
            elif kind == "done":
                stream_resp = data
                break
            elif kind == "err":
                yield f"data: {json.dumps({'type': 'error', 'message': f'Erreur recherche web : {data}'})}\n\n"
                return

        # Extraire les sources web depuis les grounding_chunks
        web_sources = []
        try:
            candidates = stream_resp.candidates if stream_resp else []
            if candidates:
                gmd = candidates[0].grounding_metadata
                if gmd:
                    for gc in gmd.grounding_chunks:
                        try:
                            if gc.HasField("web"):
                                web_sources.append({
                                    "title": gc.web.title or "",
                                    "uri": gc.web.uri or "",
                                })
                        except Exception:
                            pass
        except Exception as e:
            print(f"grounding_metadata error (non bloquant): {e}")

        done_payload = {
            "type": "done",
            "sources": [],
            "tasks_json": "",
            "reply_text": accumulated,
            "routing": "web_search",
            "web_sources": web_sources,
        }
        yield f"data: {json.dumps(done_payload)}\n\n"

        # Log usage
        try:
            usage_meta = stream_resp.usage_metadata if stream_resp else None
            in_tok = usage_meta.prompt_token_count if usage_meta else 0
            out_tok = usage_meta.candidates_token_count if usage_meta else 0
            sb.table("usage_logs").insert({
                "client_id": client_id or None,
                "user_id": user_id or None,
                "model": GEMINI_FLASH,
                "message_type": "web_search",
                "tokens_input": in_tok,
                "tokens_output": out_tok,
                "cost_usd": calculate_cost(GEMINI_FLASH, {"input_tokens": in_tok, "output_tokens": out_tok}),
            }).execute()
        except Exception as exc:
            print(f"usage_logs insert error (non bloquant): {exc}")

        # ── Langfuse generation span ───────────────────────────────────────
        if lf_trace:
            try:
                lf_trace.start_observation(
                    as_type="generation",
                    name="gemini_web_search",
                    model=GEMINI_FLASH,
                    input={"system": system[:2000], "message": message},
                    output=accumulated[:3000],
                    usage_details={"input": in_tok, "output": out_tok},
                    metadata={
                        "web_sources": [s.get("title", "") for s in web_sources],
                        "cost_usd": calculate_cost(GEMINI_FLASH, {"input_tokens": in_tok, "output_tokens": out_tok}),
                    },
                ).end()
            except Exception:
                pass

    finally:
        cancel_event.set()
        try:
            if lf_trace:
                lf_trace.end()
            _langfuse.flush()
        except Exception:
            pass


# ── chat (default) ────────────────────────────────────────────────────────────
async def chat(body: dict, user_id: Optional[str] = None):
    message = body.get("message", "")
    system = body.get("system", "")
    client_id = body.get("client_id")
    chat_history = body.get("chat_history", [])
    file_data = body.get("file")
    message_type = body.get("message_type", "chat")
    # Per-request MMR override (eval/calibration only — production omits this key)
    mmr_threshold = float(body["mmr_threshold"]) if "mmr_threshold" in body else MMR_SIM_THRESHOLD
    # Idem pour le seuil d'injection. Il porte sur le logit brut du cross-encoder,
    # une échelle qui n'a jamais été mesurée : constaté le 27/08/2026 que le top-15
    # d'une question légitime tenait entre -4.8 et -7.8, donc très en dessous du
    # seuil -2.0, et que le RAG n'injectait rien malgré un classement correct.
    # None = comportement par défaut (-1.0, ou -2.0 pour une question compte-rendu).
    inject_threshold_override = float(body["inject_threshold"]) if "inject_threshold" in body else None

    # Taille max : un prompt légitime peut atteindre ~120k chars (80k docs + tasks + context).
    # Ces limites empêchent l'abus de ressources sans bloquer aucun usage normal.
    if len(system) > 500_000 or len(message) > 20_000:
        return JSONResponse({"error": "Requête trop longue"}, status_code=400, headers=_CORS_HEADERS)

    if not client_id:
        return JSONResponse({"error": "client_id requis"}, status_code=400, headers=_CORS_HEADERS)
    await _assert_role(user_id, client_id, ["owner", "member"])

    # ── Langfuse trace (manuelle pour SSE) ─────────────────────────────────
    # v4 n'a plus `.trace()` : la trace EST son span racine, ouvert ici et fermé
    # dans le `finally` du générateur SSE (ou de _chat_web_search_stream).
    #
    # `user_id` voyage dans les métadonnées et non comme attribut de trace : en v4,
    # l'attribut de premier niveau ne se pose que via `propagate_attributes()`, un
    # context manager. Il faudrait donc ouvrir le span racine À L'INTÉRIEUR du
    # générateur, ce qui sortirait de la trace tout le pipeline RAG — construit ici,
    # avant que StreamingResponse ne soit renvoyé. Compromis assumé : on garde le RAG
    # dans la trace, on perd le filtrage par utilisateur dans l'UI Langfuse. Pour le
    # récupérer, déplacer l'ouverture du span dans `generate()` sous
    # `with propagate_attributes(user_id=user_id):` et accepter de perdre le RAG.
    _lf_trace = _langfuse.start_observation(
        name="chat",
        as_type="span",
        input=message,
        metadata={
            "client_id": client_id,
            "message_type": message_type,
            "user_id": user_id,
        },
    )

    # Web search — Gemini avec Google Search grounding, skip RAG entièrement
    if message_type == "web_search":
        return StreamingResponse(
            _chat_web_search_stream(system, message, user_id, client_id, lf_trace=_lf_trace),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    # Both task_action and chat use Gemini Flash; only generate_brief uses Pro
    chat_model = GEMINI_FLASH
    max_tokens = 1500 if message_type == "task_action" else 5000

    # RAG pipeline
    system_with_rag = system
    sources_used: list[dict] = []
    injected_chunks: list[dict] = []
    reranked: list[dict] = []
    mmr_dropped_count = 0
    _is_temporal_browse = False
    rag_error: Optional[str] = None

    # Skip RAG entirely for task actions — they only need the task list, not documents.
    if message and message_type not in ("task_action", "task_query"):
        try:
            loop = asyncio.get_running_loop()

            # HyDE (Hypothetical Document Embeddings): for substantive queries, ask
            # Gemini Flash to generate what an ideal document excerpt would look like,
            # then embed that instead of the raw question. Doc-to-doc matching is far
            # more accurate than question-to-doc for paraphrase-style models.
            # Skip for short/conversational turns — they don't benefit from it.
            query_for_embed = message

            # ── Temporal browse detection (AVANT HyDE — évite un appel Gemini inutile) ──
            # Exploration temporelle : l'utilisateur veut VOIR les docs récents,
            # pas chercher DANS les docs. Distinct de temporal_keywords qui booste
            # la pertinence temporelle dans une recherche sémantique.
            _TEMPORAL_BROWSE_PATTERNS = [
                "derniers documents", "derniers docs", "documents récents", "docs récents",
                "quoi de neuf", "nouveaux documents", "nouveaux docs", "résume les derniers",
                "résume les documents", "derniers fichiers", "fichiers récents",
                "ajoutés récemment", "modifiés récemment", "mis à jour récemment",
            ]
            _is_temporal_browse = any(p in message.lower() for p in _TEMPORAL_BROWSE_PATTERNS)

            if not _is_temporal_browse and len(message.strip()) > 25:
                try:
                    _hyde_model = genai.GenerativeModel(GEMINI_FLASH)
                    is_cr_query = any(k in message.lower() for k in CR_KEYWORDS)
                    if is_cr_query:
                        hyde_prompt = (
                            f"Écris en 2-3 phrases un extrait de compte-rendu de réunion "
                            f"professionnel qui contiendrait la réponse à cette question : "
                            f"« {message} »\n"
                            f"L'extrait doit ressembler à des notes de réunion avec des points "
                            f"d'avancement, des décisions prises, des actions à suivre.\n"
                            f"Réponds uniquement avec l'extrait, sans introduction."
                        )
                    else:
                        hyde_prompt = (
                            f"Écris en 2-3 phrases un extrait de document professionnel "
                            f"qui contiendrait la réponse à cette question : « {message} »\n"
                            f"Réponds uniquement avec l'extrait, sans introduction ni explication."
                        )
                    # generate_content() est synchrone : appelé directement, il bloquait
                    # la boucle d'événements pendant 1 à 3 s, donc toutes les requêtes
                    # servies par la même instance. On le renvoie sur le pool par défaut,
                    # comme le sont déjà l'embedding et le reranking juste en dessous.
                    _hyde_resp = await loop.run_in_executor(
                        None,
                        lambda: _hyde_model.generate_content(
                            hyde_prompt,
                            generation_config=genai.types.GenerationConfig(
                                max_output_tokens=120,
                                temperature=0.0,
                            ),
                        ),
                    )
                    _hyde_text = (_hyde_resp.text or "").strip()
                    if len(_hyde_text) > 20:
                        query_for_embed = _hyde_text
                except Exception:
                    pass  # fall back to raw message

            if _is_temporal_browse and client_id:
                def _cite_name_b(name: str) -> str:
                    return name.replace('[', '(').replace(']', ')')

                def _fmt_date(iso_str):
                    if not iso_str:
                        return "date inconnue"
                    try:
                        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
                        return dt.strftime("%d/%m/%Y")
                    except Exception:
                        return "date inconnue"

                # Étape 1 : récupérer les 10 sources les plus récentes (une ligne par source).
                # On fetche source_name + drive_modified_at sans chunk_text pour rester léger,
                # puis on déduplique en Python (PostgREST ne supporte pas DISTINCT ON).
                all_sources_result = (
                    sb.table("document_chunks")
                    .select("source_name, drive_modified_at, source_type")
                    .eq("client_id", client_id)
                    .neq("source_type", "session")
                    .eq("is_administrative", False)
                    .not_.is_("drive_modified_at", "null")
                    .order("drive_modified_at", desc=True)
                    .limit(5000)
                    .execute()
                )
                all_rows = all_sources_result.data or []

                seen_src: dict = {}
                for r in all_rows:
                    src = r["source_name"]
                    if src not in seen_src:
                        seen_src[src] = {
                            "source_name": src,
                            "drive_modified_at": r["drive_modified_at"],
                            "source_type": r["source_type"],
                        }
                    if len(seen_src) >= 10:
                        break

                top_sources = list(seen_src.values())

                # Étape 2 : récupérer le premier chunk de chaque source (10 requêtes ciblées).
                temporal_docs = []
                for src_info in top_sources:
                    chunk_result = (
                        sb.table("document_chunks")
                        .select("source_name, chunk_text, source_type, drive_modified_at")
                        .eq("client_id", client_id)
                        .eq("source_name", src_info["source_name"])
                        .limit(1)
                        .execute()
                    )
                    if chunk_result.data:
                        temporal_docs.append(chunk_result.data[0])

                if temporal_docs:
                    doc_block = "\n\n".join(
                        f"— {_cite_name_b(c['source_name'])} (modifié le {_fmt_date(c['drive_modified_at'])})\n{c['chunk_text'][:500]}"
                        for c in temporal_docs
                    )
                    system_with_rag += (
                        "\n\n[Documents récents — triés par date de modification]\n"
                        "Ces documents sont les plus récemment modifiés dans le Drive du client. "
                        "Présente-les par ordre chronologique inversé (le plus récent en premier). "
                        "Pour chaque document, donne le titre, la date de modification, et un résumé en 1-2 phrases.\n"
                        "Présente uniquement les documents de travail ci-dessous (pas de pièces comptables).\n\n"
                        + doc_block
                    )
                    sources_used = [
                        {"source_name": c["source_name"], "source_type": c["source_type"],
                         "preview": c["chunk_text"][:120]}
                        for c in temporal_docs
                    ]
                    injected_chunks = temporal_docs
                else:
                    system_with_rag += (
                        "\n\n[Disponibilité des documents]\n"
                        "Aucun document avec date de modification trouvé. "
                        "Suggère à l'utilisateur de lancer une synchronisation Drive."
                    )

            else:
                query_emb = (await loop.run_in_executor(_EMBED_EXECUTOR, embed_texts, [query_for_embed]))[0]

                # Mots-clés ≥4 chars extraits du message brut (pas du texte HyDE).
                # Utilisés pour (1) la requête FTS et (2) le safety net post-retrieval.
                # OR-joints : websearch_to_tsquery('simple', 'A OR B') → 'A'|'B' —
                # évite qu'un stopword absent des documents bloque tout le bras FTS.
                query_words = {w.lower() for w in re.findall(r'\w{4,}', message)}
                _query_text = ' OR '.join(query_words) if query_words else None
                # Défini ici et pas plus bas : la sélection « le dernier X » l'utilise
                # avant le reranking. Le 27/08/2026, l'avoir laissé après a produit un
                # NameError sur CHAQUE requête, avalé par le try/except du pipeline —
                # côté utilisateur ça ressemblait à « aucun document pertinent ».
                query_lower = message.lower()

                result = sb.rpc("match_chunks", {
                    "query_embedding": query_emb,
                    "query_text": _query_text,
                    "match_count": 30,
                    "p_client_id": client_id,
                }).execute()
                chunks = result.data or []
                # Normalize: RPC aliases source_name→source_file, chunk_text→content.
                # rrf_score remplace similarity — non consommé ici (le reranker rescores).
                chunks = [
                    {
                        **c,
                        "source_name": c.get("source_file") or c.get("source_name") or "",
                        "chunk_text": c.get("content") or c.get("chunk_text") or "",
                        "source_type": c.get("source_type") if c.get("source_type") is not None else "doc",
                        "metadata": c.get("metadata"),
                        "embedding": c.get("embedding"),
                    }
                    for c in chunks
                ]

                # Keyword-source safety net: if a source whose name matches the query
                # isn't in the semantic results (scored below rank 30), fetch its chunks
                # directly and add them to the reranker pool.
                if query_words:
                    semantic_sources = {c["source_name"] for c in chunks}
                    try:
                        all_sources = _client_source_names(client_id)
                        missing_kw_sources = [
                            s for s in all_sources
                            if s not in semantic_sources
                            and sum(1 for w in query_words if w in s.lower()) >= 2
                        ]
                        for src in missing_kw_sources[:3]:
                            # is_administrative filtré ici aussi : sans ça le filet
                            # rouvrait au RAG les pièces comptables que match_chunks
                            # exclut, en contournant la colonne.
                            extra = (
                                sb.table("document_chunks")
                                .select("source_name, chunk_text, source_type, client_id, drive_modified_at, embedding")
                                .eq("client_id", client_id)
                                .eq("source_name", src)
                                .eq("is_administrative", False)
                                .limit(2)
                                .execute()
                            )
                            for row in (extra.data or []):
                                chunks.append({
                                    **row,
                                    "rrf_score": 0.0,
                                })
                    except Exception as _kw_err:
                        print(f"keyword safety net error (non bloquant): {_kw_err}")

                # ── « le dernier X » — série datée, sélection par date ────────────────
                # « que s'est-il passé lors du DERNIER point de tracking ? » est une
                # question d'ordre chronologique, pas de similarité sémantique. Le
                # cross-encoder ne peut pas y répondre : il note la proximité d'un
                # texte à une question, il n'ordonne pas des documents entre eux.
                #
                # Mesuré en production le 27/08/2026 : les 6 notes de suivi obtenaient
                # entre -1.6 et -4.8, et le seuil d'injection (-2.0) ne laissait passer
                # que celle du 22/05 — la troisième plus ancienne. Le chat répondait donc
                # « le dernier point, daté du 22/05/2026 », faux, avec aplomb. Baisser le
                # seuil n'est pas la réponse : mesuré, il sépare correctement les cas où
                # il faut répondre de ceux où il faut s'abstenir.
                #
                # Ici on fait ce qu'un humain fait : parmi les sources dont le NOM recoupe
                # la question et qui portent une date, prendre la plus récente, et lui
                # garantir sa place dans le prompt indépendamment de son score.
                # Même logique que _TEMPORAL_BROWSE_PATTERNS, mais pour une série
                # identifiée par son sujet au lieu de tout le Drive.
                _series_source = None
                _series_date = None
                if query_words and re.search(r"\bderni[eè]re?s?\b", query_lower):
                    try:
                        _dated_rows = (
                            sb.table("document_chunks")
                            .select("source_name, doc_date")
                            .eq("client_id", client_id)
                            .eq("is_administrative", False)
                            .neq("source_type", "session")
                            .not_.is_("doc_date", "null")
                            .order("doc_date", desc=True)
                            .limit(2000)
                            .execute()
                        ).data or []
                        # Déjà trié par doc_date décroissant : le premier nom qui recoupe
                        # la question sur au moins 2 mots est le plus récent de sa série.
                        for _row in _dated_rows:
                            _name = _row.get("source_name") or ""
                            if sum(1 for w in query_words if w in _name.lower()) >= 2:
                                _series_source, _series_date = _name, _row.get("doc_date")
                                break
                    except Exception as _ser_err:
                        print(f"série datée (non bloquant): {_ser_err}")

                if _series_source and _series_source not in {c["source_name"] for c in chunks}:
                    try:
                        _extra = (
                            sb.table("document_chunks")
                            .select("source_name, chunk_text, source_type, client_id, "
                                    "drive_modified_at, doc_date, embedding")
                            .eq("client_id", client_id)
                            .eq("source_name", _series_source)
                            .limit(4)
                            .execute()
                        )
                        for _row in (_extra.data or []):
                            chunks.append({**_row, "rrf_score": 0.0})
                    except Exception as _ser_err:
                        print(f"série datée, fetch (non bloquant): {_ser_err}")

                # Cross-encoder reranking on the full pool (semantic top-30 + keyword safety
                # net chunks). Diversity cap (2 per source) applied AFTER reranking so the
                # reranker score, not insertion order, determines which chunk of a source wins.
                if RAG_DEBUG:
                    _n_safety = sum(1 for c in chunks if c.get("rrf_score") == 0.0)
                    _n_dated = sum(1 for c in chunks if c.get("drive_modified_at"))
                    _rag_log(f"question : {message[:120]!r}")
                    _rag_log(f"embed sur : {'HyDE' if query_for_embed != message else 'question brute'}")
                    _rag_log(f"fts       : {_query_text!r}")
                    _rag_log(
                        f"pool      : {len(chunks)} chunks "
                        f"({len(chunks) - _n_safety} du RPC, {_n_safety} du filet mots-clés) — "
                        f"{_n_dated} avec une date"
                    )
                    if _n_dated == 0 and chunks:
                        _rag_log("  ⚠ aucun chunk daté : la migration 20260827 n'est pas appliquée, "
                                 "le score temporel retombera sur 0.5 pour tout le monde")

                reranked = await loop.run_in_executor(_RERANK_EXECUTOR, _rerank_chunks, message, chunks)

                temporal_keywords = {"récent", "dernier", "actuelle", "actuel", "aujourd", "maintenant", "nouveau", "dernière"}
                decay = 30 if any(k in query_lower for k in temporal_keywords) else 180

                for c in reranked:
                    # Les chunks `session` sont des résumés générés par le modèle, insérés
                    # sans drive_modified_at : le RPC leur substitue created_at, donc « écrit
                    # aujourd'hui » devenait « document le plus frais » et ils raflaient le
                    # bonus temporel maximal face aux vrais documents. Score neutre : leur
                    # date d'écriture ne dit rien de la fraîcheur de l'information.
                    if c.get("source_type") == "session":
                        t = 0.5
                    else:
                        t = _temporal_score(c.get("drive_modified_at"), decay)
                    # ⚠️ NE PAS normaliser le logit par sigmoid. Tenté le 27/08/2026,
                    # annulé le jour même sur mesure en production : sigmoid écrase tout
                    # le régime négatif vers 0 (sigmoid(-9.6) ≈ 7e-5), le terme 0.7×r
                    # devient négligeable devant 0.3×t, et le classement se réduit à un
                    # tri par date de modification. Observé : un chunk à rerank -9.594
                    # classé PREMIER du pool avec final_score 0.29, uniquement parce que
                    # son fichier avait été modifié la veille.
                    #
                    # L'asymétrie d'échelle entre les deux termes (logit dans [-11, +11]
                    # contre temporel dans [0, 1]) est donc VOLONTAIRE : elle fait du
                    # score temporel un départage entre chunks de pertinence comparable,
                    # et non un critère de classement. C'est le comportement voulu.
                    c["_temporal"] = t          # conservé pour la trace uniquement
                    c["final_score"] = 0.7 * c.get("rerank_score", 0.0) + 0.3 * t

                reranked.sort(key=lambda c: c["final_score"], reverse=True)

                if RAG_DEBUG:
                    _hit = [k for k in temporal_keywords if k in query_lower]
                    _rag_log(f"decay     : {decay} j" + (f" (mot temporel : {_hit})" if _hit else " (aucun mot temporel)"))
                    _rag_log_pool("après rerank + score temporel", reranked)

                # MMR dedup: embeddings are returned by match_chunks RPC (and fetched for
                # safety net chunks), so no re-encoding needed — dot product only (~10ms).
                # Guard: if RPC not yet redeployed, embeddings are None → skip MMR silently.
                mmr_dropped_count = 0
                try:
                    mmr_embs = [c.get("embedding") for c in reranked]
                    if mmr_embs and mmr_embs[0] is not None:
                        pre_mmr_count = len(reranked)
                        reranked = _mmr_filter(reranked, mmr_embs, mmr_threshold)
                        mmr_dropped_count = pre_mmr_count - len(reranked)
                except Exception as _mmr_err:
                    print(f"MMR filter error (non bloquant, ignoré): {_mmr_err}")

                if RAG_DEBUG:
                    _rag_log(
                        f"MMR       : {mmr_dropped_count} écartés (seuil {mmr_threshold}) "
                        f"→ {len(reranked)} restants"
                    )

                safety_net_sources = {
                    c["source_name"] for c in chunks
                    if c.get("rrf_score") == 0.0
                }

                # Focused document detection: si une source domine le top-10
                # reranké (≥3 chunks), c'est une requête ciblée sur un doc
                # → on augmente son cap à 4 pour injecter plus de contenu.
                _top10_sources = [c["source_name"] for c in reranked[:10]]
                _src_counts = Counter(_top10_sources)
                _focused_source = None
                for _src, _cnt in _src_counts.most_common(1):
                    if _cnt >= 3:
                        _focused_source = _src

                seen_src: dict = {}
                diverse: list = []
                for c in reranked:
                    src = c["source_name"]
                    if src in safety_net_sources:
                        cap = 1
                    elif src == _focused_source:
                        cap = 4
                    else:
                        cap = 2
                    if seen_src.get(src, 0) < cap:
                        diverse.append(c)
                        seen_src[src] = seen_src.get(src, 0) + 1

                if diverse:
                    is_account_query = any(k in query_lower for k in CR_KEYWORDS)
                    inject_threshold = (
                        inject_threshold_override
                        if inject_threshold_override is not None
                        else (-2.0 if is_account_query else -1.0)
                    )
                    MAX_INJECT = 8 if is_account_query else 6

                    # Le seuil porte sur le logit BRUT du cross-encoder, pas sur
                    # final_score : -1.0 / -2.0 ont été calibrés sur cette échelle.
                    # final_score est désormais dans [0, 1] (sigmoid) et ne s'y
                    # compare plus — l'utiliser ici laisserait tout passer.
                    # Le plus récent de la série passe sans condition de score : c'est la
                    # date qui le qualifie, pas la pertinence sémantique de son texte.
                    #
                    # Lu sur `reranked` et non sur `diverse` : ses chunks sont ajoutés au
                    # pool avec rrf_score = 0.0, ce qui les fait entrer dans
                    # safety_net_sources, où le cap de diversité est de 1. Passer par
                    # `diverse` n'aurait donc injecté qu'UN chunk sur les quatre — le bon
                    # document, mais un quart de son contenu, ce qui ne répond pas à
                    # « que s'est-il passé lors de… ». `reranked` est déjà post-MMR, les
                    # quasi-doublons sont donc écartés.
                    series_chunks = ([c for c in reranked if c["source_name"] == _series_source][:4]
                                     if _series_source else [])
                    guaranteed = [c for c in diverse
                                  if c["source_name"] in safety_net_sources
                                  and c["source_name"] != _series_source
                                  and c.get("rerank_score", 0.0) >= inject_threshold]
                    normal = [c for c in diverse
                              if c["source_name"] not in safety_net_sources
                              and c["source_name"] != _series_source
                              and c.get("rerank_score", 0.0) >= inject_threshold]

                    to_inject = (series_chunks + guaranteed + normal)[:MAX_INJECT]
                    injected_chunks = to_inject

                    doc_chunks = [c for c in to_inject if c["source_type"] != "session"]
                    session_chunks = [c for c in to_inject if c["source_type"] == "session"]

                    if RAG_DEBUG:
                        if _series_source:
                            _rag_log(f"série     : plus récent = {_series_source!r} "
                                     f"({str(_series_date)[:10]}) — {len(series_chunks)} chunks garantis")
                        _rag_log(
                            f"injection : seuil {inject_threshold}, max {MAX_INJECT} → "
                            f"{len(to_inject)} chunks ({len(doc_chunks)} doc, {len(session_chunks)} session)"
                        )
                        _rag_log_pool("injecté dans le prompt", to_inject, limit=MAX_INJECT)
                        # Comparaison par identité : deux chunks distincts peuvent avoir
                        # un contenu identique, `in` sur des dicts les confondrait.
                        _kept_ids = {id(c) for c in to_inject}
                        _rejected = [c for c in diverse if id(c) not in _kept_ids]
                        if _rejected:
                            _rag_log_pool("écarté (sous le seuil, ou au-delà du max)", _rejected, limit=10)

                    if doc_chunks:
                        # Normalize source names for in-text citation: replace [ ] with ( )
                        # so filenames like "[Client x Agency] || Doc" don't produce nested
                        # brackets [[Client x Agency] || Doc] that break the citation regex.
                        def _cite_name(name: str) -> str:
                            return name.replace('[', '(').replace(']', ')')

                        doc_block = "\n\n".join(
                            f"— {_cite_name(c['source_name'])}\n{c['chunk_text']}" for c in doc_chunks
                        )
                        _series_hint = ""
                        if _series_source and any(c["source_name"] == _series_source for c in doc_chunks):
                            _series_hint = (
                                f"\nLe document « {_cite_name(_series_source)} » est, parmi ceux dont le nom "
                                f"correspond à la question, celui dont la date est la plus récente"
                                + (f" ({_series_date[:10]})" if _series_date else "")
                                + ". Appuie-toi sur lui pour répondre à « le dernier ».\n"
                            )
                        system_with_rag += (
                            "\n\n[Documents pertinents]\n"
                            + _series_hint +
                            "RÈGLE DE FIABILITÉ (impérative) :\n"
                            "- Réponds en t'appuyant UNIQUEMENT sur les extraits ci-dessous, la fiche client "
                            "et l'historique fourni.\n"
                            "- N'invente rien et ne comble pas les trous avec des suppositions ou des "
                            "connaissances générales sur le sujet.\n"
                            "- Si les extraits ne répondent que partiellement, distingue clairement ce qui est "
                            "étayé par les documents de ce qui ne l'est pas, et dis explicitement ce que tu ne "
                            "trouves pas dans les sources.\n"
                            "- Quand tu utilises une information issue d'un extrait, cite le nom du fichier "
                            "source entre crochets, ex : [NomDuFichier].\n"
                            "- Les extraits sont une SÉLECTION, pas un inventaire. N'en déduis jamais qu'un "
                            "document est le plus récent, le premier ou le seul de sa série au motif qu'il "
                            "t'a été fourni. Si la question porte sur « le dernier » quelque chose et que "
                            "tu ne vois qu'un document daté, dis de quand il date et précise que tu ne peux "
                            "pas garantir qu'il s'agit du plus récent.\n\n"
                            + doc_block
                        )
                        sources_used = [
                            {
                                "source_name": c["source_name"],
                                "source_type": c["source_type"],
                                "preview": c["chunk_text"][:120],
                            }
                            for c in doc_chunks
                        ]

                    if session_chunks:
                        sess_block = "\n\n".join(
                            f"— {c['source_name']}\n{c['chunk_text']}" for c in session_chunks
                        )
                        system_with_rag += (
                            "\n\n[Historique pertinent — NON VÉRIFIÉ]\n"
                            "Ces extraits sont des résumés de tes propres conversations passées, "
                            "rédigés automatiquement. Ce ne sont PAS des sources : ils peuvent "
                            "contenir des erreurs, y compris des dates ou des chiffres que tu as "
                            "inventés lors d'un échange précédent.\n"
                            "- Ne t'en sers que pour te rappeler le fil des échanges.\n"
                            "- N'en tire JAMAIS un fait, une date, un chiffre ni un nom que les "
                            "documents ci-dessus ne confirment pas.\n"
                            "- Si l'information ne se trouve que là, dis qu'elle vient d'un échange "
                            "précédent et qu'elle n'est pas étayée par les documents.\n"
                            "- Ne les cite pas entre crochets — ce ne sont pas des fichiers sources."
                            "\n\n" + sess_block
                        )

                    if not doc_chunks:
                        # Aucun extrait documentaire retenu. Le cas `session_chunks` non
                        # vide était le plus dangereux : le modèle ne recevait qu'un résumé
                        # de ses propres réponses et le resservait comme un fait sourcé.
                        # C'est ce qui a produit le « point de tracking du 29 juin 2026 »,
                        # date absente de tous les documents Drive (27/08/2026).
                        system_with_rag += (
                            "\n\n[Disponibilité des documents]\n"
                            "Aucun extrait pertinent trouvé dans les documents indexés pour cette question. "
                            "Si tu ne trouves pas l'information dans la fiche client ou le contexte disponible, "
                            "dis-le explicitement plutôt que d'estimer ou d'inventer. Ne réponds PAS depuis des "
                            "connaissances générales comme s'il s'agissait d'informations vérifiées sur ce client — "
                            "propose plutôt à l'utilisateur de préciser ou d'indiquer le document concerné."
                        )
                else:
                    # No documents indexed for this client at all
                    system_with_rag += (
                        "\n\n[Disponibilité des documents]\n"
                        "Aucun extrait pertinent trouvé dans les documents indexés pour cette question. "
                        "Si tu ne trouves pas l'information dans la fiche client ou le contexte disponible, "
                        "dis-le explicitement plutôt que d'estimer ou d'inventer. Ne réponds PAS depuis des "
                        "connaissances générales comme s'il s'agissait d'informations vérifiées sur ce client — "
                        "propose plutôt à l'utilisateur de préciser ou d'indiquer le document concerné."
                    )
        except Exception as e:
            # Ce except large transforme n'importe quelle erreur de code en « aucun
            # document pertinent » côté utilisateur — un NameError est passé comme ça
            # en production le 27/08/2026. On garde le comportement non bloquant, mais
            # l'erreur devient visible dans le payload debug et dans la trace.
            rag_error = f"{type(e).__name__}: {e}"
            print(f"RAG pipeline error (non bloquant): {e}")
            print(traceback.format_exc())
            _rag_log(f"⚠ PIPELINE EN ERREUR — {rag_error}")

    # Build history for Gemini — "u"→"user", "a"→"model"; must start with user
    raw_hist = list(chat_history)
    while raw_hist and raw_hist[0]["role"] == "a":
        raw_hist.pop(0)

    history_contents = [
        {"role": "user" if m["role"] == "u" else "model", "parts": [m["text"]]}
        for m in raw_hist
    ]

    # Build current user message parts — multimodal if file attached
    system_final = system_with_rag
    if file_data and file_data.get("data") and file_data.get("mediaType"):
        allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/gif", "image/webp"]
        if file_data["mediaType"] not in allowed_types:
            return JSONResponse(
                {"error": f"Type de fichier non supporté : {file_data['mediaType']}"},
                status_code=400,
            )
        system_final += (
            "\n\nL'utilisateur t'a partagé un fichier. Extrais les informations clés : "
            "type de document, points importants, données chiffrées, actions suggérées."
        )
        current_parts = [
            {"inline_data": {"mime_type": file_data["mediaType"], "data": file_data["data"]}},
            message or "Analyse ce fichier.",
        ]
    else:
        current_parts = [message]

    contents = history_contents + [{"role": "user", "parts": current_parts}]

    # ── Streaming SSE — tokens envoyés au fur et à mesure ────────────────────
    loop = asyncio.get_running_loop()

    async def generate():
        q: asyncio.Queue = asyncio.Queue()
        # cancel_event : positionné à True quand le générateur est abandonné
        # (déconnexion client → FastAPI appelle aclose() → finally ci-dessous).
        # Le thread _sync_stream le vérifie entre chaque chunk Gemini pour s'arrêter
        # proprement sans laisser la connexion Gemini ouverte inutilement.
        cancel_event = threading.Event()

        def _sync_stream():
            try:
                gm = genai.GenerativeModel(
                    model_name=chat_model,
                    system_instruction=system_final,
                    generation_config={"max_output_tokens": max_tokens},
                    safety_settings=_SAFETY_OFF,
                )
                resp = gm.generate_content(contents, stream=True,
                                           request_options={"timeout": 120})
                for chunk in resp:
                    if cancel_event.is_set():
                        break  # Déconnexion détectée — on arrête de lire Gemini
                    try:
                        t = chunk.text
                        if t:
                            loop.call_soon_threadsafe(q.put_nowait, ("tok", t))
                    except Exception:
                        pass
                if not cancel_event.is_set():
                    loop.call_soon_threadsafe(q.put_nowait, ("done", resp))
            except Exception as exc:
                if not cancel_event.is_set():
                    loop.call_soon_threadsafe(q.put_nowait, ("err", str(exc)))

        threading.Thread(target=_sync_stream, daemon=True).start()

        try:
            accumulated = ""
            stream_resp = None

            while True:
                try:
                    kind, data = await asyncio.wait_for(q.get(), timeout=120.0)
                except asyncio.TimeoutError:
                    yield f"data: {json.dumps({'type': 'error', 'message': 'Timeout IA'})}\n\n"
                    return
                if kind == "tok":
                    accumulated += data
                    yield f"data: {json.dumps({'type': 'token', 'text': data})}\n\n"
                elif kind == "done":
                    stream_resp = data
                    break
                elif kind == "err":
                    print(f"chat stream error: {data}")
                    yield f"data: {json.dumps({'type': 'error', 'message': f'Erreur IA : {data}'})}\n\n"
                    return

            # Séparer le texte affiché et le JSON de tâches
            parts = accumulated.split("---JSON---")
            reply_text = parts[0].strip()
            tasks_json = parts[1].strip() if len(parts) > 1 else ""

            debug_info = None
            if body.get("debug"):
                injected_names = {s["source_name"] for s in sources_used}
                if rag_error:
                    debug_info = [{"rag_error": rag_error}]
                elif _is_temporal_browse:
                    debug_info = [{"temporal_browse": True, "docs_found": len(injected_chunks)}]
                else:
                    debug_info = [
                        {
                            "source_name": c["source_name"],
                            "rerank_score": round(c.get("rerank_score", 0.0), 3),
                            "final_score": round(c.get("final_score", 0.0), 3),
                            "injected": c["source_name"] in injected_names,
                            "preview": c["chunk_text"][:80],
                        }
                        for c in reranked[:15]
                    ]
                    debug_info.insert(0, {
                        "mmr_threshold": mmr_threshold,
                        "mmr_dropped": mmr_dropped_count,
                    })

            done_payload = {"type": "done", "sources": sources_used, "tasks_json": tasks_json, "reply_text": reply_text, "routing": message_type}
            if debug_info is not None:
                done_payload["debug"] = debug_info
                done_payload["injected_context"] = "\n\n".join(
                    f"[{c['source_name']}]\n{c['chunk_text']}" for c in injected_chunks
                )
            yield f"data: {json.dumps(done_payload)}\n\n"

            # Log usage — non-bloquant
            try:
                usage_meta = stream_resp.usage_metadata if stream_resp else None
                in_tok = usage_meta.prompt_token_count if usage_meta else 0
                out_tok = usage_meta.candidates_token_count if usage_meta else 0
                log_row: dict = {
                    "client_id": client_id or None,
                    "model": chat_model,
                    "message_type": message_type,
                    "tokens_input": in_tok,
                    "tokens_output": out_tok,
                    "cost_usd": calculate_cost(chat_model, {"input_tokens": in_tok, "output_tokens": out_tok}),
                }
                if user_id:
                    log_row["user_id"] = user_id
                sb.table("usage_logs").insert(log_row).execute()
            except Exception as exc:
                print(f"usage_logs insert error (non bloquant): {exc}")

            # ── Langfuse generation span ───────────────────────────────────
            try:
                _lf_trace.start_observation(
                    as_type="generation",
                    name="gemini_stream",
                    model=chat_model,
                    input={"system": system_final[:3000], "message": message},
                    output=reply_text[:3000],
                    usage_details={"input": in_tok, "output": out_tok},
                    metadata={
                        "sources": [s["source_name"] for s in sources_used],
                        "message_type": message_type,
                        "temporal_browse": _is_temporal_browse,
                        "cost_usd": calculate_cost(chat_model, {"input_tokens": in_tok, "output_tokens": out_tok}),
                    },
                ).end()
                _lf_trace.update(output=reply_text[:3000])
            except Exception as _lf_err:
                print(f"langfuse chat error (non bloquant): {_lf_err}")

        finally:
            # Garantit que le thread s'arrête même si le client se déconnecte
            # en plein stream (GeneratorExit levé par FastAPI → finally déclenché).
            cancel_event.set()
            try:
                _lf_trace.end()      # v4 : la trace reste ouverte sans ça
                _langfuse.flush()
            except Exception:
                pass

    return StreamingResponse(generate(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })
